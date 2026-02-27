"""
Excel sheet parser for SOLARA - Daily Sales Tracking FY 25-26.xlsx.

Handles ALL cross-month column inconsistencies found across 101 sheets:
  - Blinkit extra DRR column (Apr–Jul 2025)
  - Flipkart missing DOC/MTD DRR (Apr–Jun 2025)
  - Amazon col[1] backtick bug (Oct 2025)
  - Myntra SKU at col[3] not col[0]
  - Leading/trailing spaces in sheet names
  - SKU codes with embedded newlines (Flipkart Oct/Nov)
  - #DIV/0! values in DOC/DRR columns
  - Future-date NaN placeholder cells
"""
import re
import logging
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any

import numpy as np
import pandas as pd

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Sheet name → portal slug
# ---------------------------------------------------------------------------

PORTAL_PATTERNS: list[tuple[str, str]] = [
    # Order matters: more specific first
    (r"^AZ IN .+ Summary", "az_summary"),       # skip — category-level
    (r"^AZ IN .+ Combo",   "amazon"),
    (r"^AZ IN ",           "amazon"),
    (r"Zepto",             "zepto"),
    (r"Swiggy",            "swiggy"),
    (r"Blinkit",           "blinkit"),
    (r"Myntra",            "myntra"),
    (r"Flipkart",          "flipkart"),
    (r"Shopify .+ Combo",  "shopify"),
    (r"Shopify",           "shopify"),
    (r"Summary",           "summary"),           # monthly summary — skip
]

SKIP_PORTALS = {"az_summary", "summary"}


def sheet_to_portal(sheet_name: str) -> str | None:
    """
    Map a raw sheet name to a portal slug.
    Returns None for sheets that should be skipped.
    """
    name = sheet_name.strip()
    for pattern, slug in PORTAL_PATTERNS:
        if re.search(pattern, name, re.IGNORECASE):
            return None if slug in SKIP_PORTALS else slug
    return None


def sheet_to_year_month(sheet_name: str) -> tuple[int, int] | None:
    """
    Extract (year, month) from a sheet name like 'Zepto FEB-26' or 'AZ IN AUG-25'.
    Returns None if not parseable.
    """
    MONTH_MAP = {
        "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
        "june": 6, "jul": 7, "july": 7, "aug": 8, "sep": 9, "oct": 10,
        "nov": 11, "dec": 12, "january": 1, "february": 2, "march": 3,
        "april": 4, "august": 8, "september": 9, "october": 10,
        "november": 11, "december": 12,
    }
    name = sheet_name.strip().lower()
    # Try "MMM-YY" pattern (e.g., feb-26, aug-25)
    m = re.search(r"(\w+)[- ](\d{2})$", name)
    if m:
        mon_str, yr_str = m.group(1), m.group(2)
        mon = MONTH_MAP.get(mon_str)
        if mon:
            year = 2000 + int(yr_str)
            return year, mon
    # Try "Month-Year" in summary sheets (e.g., "February Summary")
    for mon_str, mon_num in MONTH_MAP.items():
        if mon_str in name:
            yr_m = re.search(r"20(\d{2})", name)
            year = 2000 + int(yr_m.group(1)) if yr_m else 2025
            return year, mon_num
    return None


# ---------------------------------------------------------------------------
# Column configuration per portal
# ---------------------------------------------------------------------------

@dataclass
class PortalColMap:
    """Column index mapping for a portal sheet."""
    sku_col: int          # column that holds the SOL- SKU code
    portal_id_col: int    # column that holds portal-specific ID (ASIN, Style ID, etc.)
    name_col: int         # column for product name/title
    category_col: int     # L2 product category
    asp_col: int          # BAU ASP / BAU Price
    units_col: int        # MTD Units
    revenue_col: int | None = None   # MTD Value / MTD Revenue (None = calculate from units×asp)
    inv_cols: dict = field(default_factory=dict)  # {'portal': col_idx, 'backend': ..., ...}
    doc_col: int | None = None
    open_po_col: int | None = None
    # Amazon-specific
    target_units_col: int | None = None
    target_revenue_col: int | None = None
    target_drr_col: int | None = None
    achievement_col: int | None = None


# Base column maps — adjusted dynamically for historical schema variations
BASE_COL_MAPS: dict[str, PortalColMap] = {
    "zepto": PortalColMap(
        sku_col=0, portal_id_col=3, name_col=2, category_col=1,
        asp_col=5, units_col=4, revenue_col=7,
        inv_cols={"portal": 8},
    ),
    "swiggy": PortalColMap(
        sku_col=0, portal_id_col=3, name_col=2, category_col=1,
        asp_col=5, units_col=4, revenue_col=6,
        inv_cols={"portal": 7},
    ),
    "blinkit": PortalColMap(
        sku_col=1, portal_id_col=4, name_col=3, category_col=2,
        asp_col=10, units_col=8, revenue_col=9,
        inv_cols={"backend": 5, "frontend": 6, "solara": 7},
    ),
    "myntra": PortalColMap(
        sku_col=3, portal_id_col=0, name_col=1, category_col=2,
        asp_col=6, units_col=5, revenue_col=None,
        inv_cols={"portal": 4},
    ),
    "flipkart": PortalColMap(
        sku_col=0, portal_id_col=3, name_col=2, category_col=1,
        asp_col=8, units_col=7, revenue_col=None,
        inv_cols={"portal": 4},
        doc_col=5,
    ),
    "amazon": PortalColMap(
        sku_col=1, portal_id_col=4, name_col=3, category_col=2,
        asp_col=9, units_col=12, revenue_col=14,
        inv_cols={"solara": 15, "amazon_fc": 16},
        doc_col=17, open_po_col=18,
        target_units_col=6, target_revenue_col=7, target_drr_col=8,
        achievement_col=5,
    ),
    "shopify": PortalColMap(
        sku_col=1, portal_id_col=1, name_col=2, category_col=0,
        asp_col=5, units_col=4, revenue_col=None,
        inv_cols={},
    ),
}


def _adapt_col_map(portal: str, df: pd.DataFrame, year: int, month: int) -> PortalColMap:
    """
    Dynamically adjust column positions based on cross-month schema variations.
    Returns a copy of the base map with corrections applied.
    """
    import copy
    cm = copy.deepcopy(BASE_COL_MAPS[portal])
    cols = list(df.columns)
    non_date_count = sum(1 for c in cols if not isinstance(c, datetime))

    if portal == "blinkit":
        # Apr–Jul 2025: extra 'DRR' col at position 8 shifts everything by +1
        # Detect: check if col[8] header text contains 'DRR' (not a date)
        if len(cols) > 8:
            col8_str = str(cols[8]).strip().upper()
            if "DRR" in col8_str and not isinstance(cols[8], datetime):
                cm.units_col = 9
                cm.revenue_col = 10
                cm.asp_col = 11
                # date columns start 1 later too (handled in date detection)

    elif portal == "flipkart":
        # Apr–Jun 2025: no DOC (col 5) or MTD DRR (col 6); dates start at col 7 → col 5 early
        # Detect: non-date cols before first date col < 9
        first_date_idx = next((i for i, c in enumerate(cols) if isinstance(c, datetime)), None)
        if first_date_idx is not None and first_date_idx <= 6:
            # Old schema: E=VC Inv, F=MTD Units, G=BAU ASP, dates from H
            cm.asp_col = 6
            cm.units_col = 5
            cm.revenue_col = None
            cm.doc_col = None

    elif portal == "amazon":
        # Oct-25: col[1] header is backtick '`' instead of 'SKU ID' — position still correct
        # No adjustment needed — we use positional access anyway
        # Dec-25: col N header is 'Inventory (Months_' — still position 13, fine
        pass

    return cm


# ---------------------------------------------------------------------------
# Safe value extraction helpers
# ---------------------------------------------------------------------------

def _float(val: Any, default: float | None = None) -> float | None:
    if val is None:
        return default
    s = str(val).strip().replace(",", "").replace("₹", "").replace("#DIV/0!", "").strip()
    if s in ("", "nan", "None", "NaT", "#DIV/0!"):
        return default
    try:
        return float(s)
    except (ValueError, TypeError):
        return default


def clean_sku(val: Any) -> str | None:
    """Strip spaces, newlines; return None if not a SOL- code."""
    if val is None:
        return None
    s = str(val).strip().strip("\n").strip("\r").strip()
    return s if s.upper().startswith("SOL-") else None


# ---------------------------------------------------------------------------
# Core sheet reader
# ---------------------------------------------------------------------------

@dataclass
class SheetData:
    sheet_name: str
    portal: str
    year: int
    month: int
    col_map: PortalColMap
    sku_rows: pd.DataFrame          # filtered: only SOL- product rows, no totals
    date_columns: list[tuple[int, date]]  # [(col_index, date_value), ...]
    last_data_date: date | None     # last date with at least one non-NaN/non-zero value
    ad_spend_row: pd.Series | None  # the "Total Ad Spend" row if found


def _parse_sheet_df(
    df: pd.DataFrame, sheet_name: str, portal: str, year: int, month: int
) -> SheetData | None:
    """
    Core sheet parser — operates on an already-loaded DataFrame.
    Called by both read_sheet() (pd.ExcelFile path) and iter_sheets_ro() (openpyxl path).
    """
    cols = list(df.columns)

    # Adapt column map for historical schema differences
    cm = _adapt_col_map(portal, df, year, month)

    # ----- Find date columns -----
    date_columns: list[tuple[int, date]] = []
    for i, col in enumerate(cols):
        if isinstance(col, datetime):
            date_columns.append((i, col.date()))

    # ----- Find SKU rows -----
    sku_col_idx = cm.sku_col

    def is_sku(val) -> bool:
        return clean_sku(val) is not None

    def is_stop_row(row) -> bool:
        """True if any non-date cell in row contains 'Total Revenue' — marks end of SKU section."""
        for i, val in enumerate(row):
            if not isinstance(cols[i], datetime):
                if "total revenue" in str(val).lower():
                    return True
        return False

    sku_mask = []
    stop_found = False
    for _, row in df.iterrows():
        if stop_found:
            sku_mask.append(False)
            continue
        if is_stop_row(row):
            stop_found = True
            sku_mask.append(False)
            continue
        sku_mask.append(is_sku(row.iloc[sku_col_idx] if sku_col_idx < len(row) else None))

    sku_rows = df[sku_mask].copy()

    # ----- Find ad spend row -----
    ad_spend_row = None
    try:
        all_rows = list(df.iterrows())
        for _, row in all_rows:
            row_vals = [str(v).lower() for v in row if str(v).strip() not in ("", "nan", "None")]
            if any("total ad spend" in v for v in row_vals):
                ad_spend_row = row
                break
    except Exception:
        pass

    # ----- Find last date with actual data -----
    last_data_date: date | None = None
    today = datetime.today().date()
    for col_idx, col_date in reversed(date_columns):
        if col_date > today:
            continue
        col_vals = sku_rows.iloc[:, col_idx].apply(
            lambda v: _float(v) if _float(v) is not None else None
        )
        if col_vals.notna().any() and (col_vals.dropna() != 0).any():
            last_data_date = col_date
            break

    return SheetData(
        sheet_name=sheet_name,
        portal=portal,
        year=year,
        month=month,
        col_map=cm,
        sku_rows=sku_rows,
        date_columns=date_columns,
        last_data_date=last_data_date,
        ad_spend_row=ad_spend_row,
    )


def read_sheet(xl: pd.ExcelFile, sheet_name: str) -> SheetData | None:
    """
    Parse one portal sheet and return a SheetData.
    Returns None for sheets that should be skipped.
    """
    portal = sheet_to_portal(sheet_name)
    if portal is None:
        logger.debug(f"Skipping sheet: {sheet_name!r}")
        return None

    period = sheet_to_year_month(sheet_name)
    if period is None:
        logger.warning(f"Could not parse year/month from sheet: {sheet_name!r}")
        return None

    year, month = period

    try:
        df = pd.read_excel(xl, sheet_name=sheet_name, header=1, dtype=object)
    except Exception as e:
        logger.error(f"Failed to read sheet {sheet_name!r}: {e}")
        return None

    return _parse_sheet_df(df, sheet_name, portal, year, month)


def iter_sheets(xl: pd.ExcelFile) -> list[SheetData]:
    """
    Iterate all sheets in the workbook and return parsed SheetData objects.
    Skips summary sheets and sheets with errors.
    """
    results = []
    for sheet_name in xl.sheet_names:
        sd = read_sheet(xl, sheet_name)
        if sd is not None:
            results.append(sd)
    return results


def _ws_to_df(ws) -> pd.DataFrame:
    """
    Convert an openpyxl worksheet to a pandas DataFrame equivalent to
    pd.read_excel(xl, sheet_name=..., header=1, dtype=object).

    header=1: second row (index 1) is used as column headers;
              first row (index 0) is skipped, rows 2+ are data.
    Empty cells come through as None (vs NaN from pd.read_excel) — both
    are handled correctly by _float() and clean_sku().

    Uses ws.max_row / ws.max_column (from the sheet's declared dimensions)
    to avoid iterating past the last used row/column.  Also stops early
    after the 'Total Revenue' sentinel row — everything below is summary
    data, not SKU rows (same marker _parse_sheet_df uses).
    """
    try:
        max_row = getattr(ws, "max_row", None)
        max_col = getattr(ws, "max_column", None)
        rows: list[list] = []
        for row in ws.iter_rows(values_only=True, max_row=max_row, max_col=max_col):
            rows.append(list(row))
            # Skip stop-check on title (row 0) and header (row 1)
            if len(rows) > 2 and any(
                v is not None and "total revenue" in str(v).lower() for v in row
            ):
                break
    except Exception:
        return pd.DataFrame()
    if len(rows) < 2:
        return pd.DataFrame()
    headers = rows[1]          # row index 1 → header row
    data = rows[2:]            # rows 2+ → data
    return pd.DataFrame(data, columns=headers)


def iter_sheets_ro(path: str) -> list[SheetData]:
    """
    Faster alternative to iter_sheets() that opens the workbook with
    openpyxl read_only=True + data_only=True.

    Skips loading styles, conditional formatting, and data validation XML —
    typically 3–5x faster than the standard pd.ExcelFile path for large
    workbooks (100+ sheets).  Reads ALL portal data; only non-portal tabs
    (Summary, etc.) are skipped, same as iter_sheets().
    """
    import warnings
    import openpyxl

    results = []
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet_name in wb.sheetnames:
            portal = sheet_to_portal(sheet_name)
            if portal is None:
                continue
            period = sheet_to_year_month(sheet_name)
            if period is None:
                logger.warning(f"Could not parse year/month from sheet: {sheet_name!r}")
                continue
            year, month = period
            ws = wb[sheet_name]
            try:
                df = _ws_to_df(ws)
                if df.empty:
                    continue
                sd = _parse_sheet_df(df, sheet_name, portal, year, month)
                if sd is not None:
                    results.append(sd)
            except Exception as e:
                logger.error(f"Failed to read sheet {sheet_name!r}: {e}")
    finally:
        wb.close()
    return results


def get_snapshot_date(sd: SheetData) -> date:
    """
    Determine the inventory snapshot date for a sheet:
    - For the current month: last date with non-NaN/non-zero data
    - For historical months: last day of the month
    """
    import calendar
    today = datetime.today().date()
    last_day = date(sd.year, sd.month, calendar.monthrange(sd.year, sd.month)[1])

    if sd.year == today.year and sd.month == today.month:
        return sd.last_data_date or last_day
    return last_day

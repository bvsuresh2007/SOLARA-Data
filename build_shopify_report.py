import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import Counter

RAW = """SOL1194557|SHP27-00671|To Deliver and Bill|No DN|||
SOL1194555|SHP27-00669|To Deliver and Bill|No DN|||
SOL1194554|SHP27-00668|To Deliver and Bill|No DN|||
SOL1194553|SHP27-00667|To Deliver and Bill|No DN|||
SOL1194551|SHP27-00665|To Deliver and Bill|No DN|||
SOL1194549|SHP27-00663|To Deliver and Bill|No DN|||
SOL1194548|SHP27-00662|To Deliver and Bill|No DN|||
SOL1194546|SHP27-00660|To Deliver and Bill|No DN|||
SOL1194544|SHP27-00658|To Deliver and Bill|No DN|||
SOL1194543|SHP27-00657|To Deliver and Bill|No DN|||
SOL1194542|SHP27-00656|To Deliver and Bill|No DN|||
SOL1194541|SHP27-00655|To Deliver and Bill|No DN|||
SOL1194539|SHP27-00653|To Deliver and Bill|No DN|||
SOL1194537|SHP27-00651|To Deliver and Bill|No DN|||
SOL1194536|SHP27-00650|To Deliver and Bill|No DN|||
SOL1194535|SHP27-00649|To Deliver and Bill|No DN|||
SOL1194534|SHP27-00648|To Deliver and Bill|No DN|||
SOL1194533|SHP27-00647|To Deliver and Bill|No DN|||
SOL1194530|SHP27-00644|To Deliver and Bill|No DN|||
SOL1194529|SHP27-00643|To Deliver and Bill|No DN|||
SOL1194528|SHP27-00642|To Deliver and Bill|No DN|||
SOL1194527|SHP27-00641|To Deliver and Bill|No DN|||
SOL1194525|SHP27-00639|To Deliver and Bill|No DN|||
SOL1194523|SHP27-00637|To Deliver and Bill|No DN|||
SOL1194522|SHP27-00636|To Deliver and Bill|No DN|||
SOL1194519|SHP27-00633|To Deliver and Bill|No DN|||
SOL1194517|SHP27-00631|To Deliver and Bill|No DN|||
SOL1194516|SHP27-00630|To Deliver and Bill|No DN|||
SOL1194515|SHP27-00629|To Deliver and Bill|No DN|||
SOL1194514|SHP27-00628|To Deliver and Bill|No DN|||
SOL1194513|SHP27-00627|To Deliver and Bill|No DN|||
SOL1194512|SHP27-00626|To Deliver and Bill|No DN|||
SOL1194511|SHP27-00625|To Deliver and Bill|No DN|||
SOL1194510|SHP27-00624|To Deliver and Bill|No DN|||
SOL1194509|SHP27-00623|To Deliver and Bill|No DN|||
SOL1194508|SHP27-00622|To Deliver and Bill|No DN|||
SOL1194507|SHP27-00621|To Deliver and Bill|No DN|||
SOL1194506|SHP27-00620|To Deliver and Bill|No DN|||
SOL1194505|SHP27-00619|To Deliver and Bill|No DN|||
SOL1194502|SHP27-00616|To Deliver and Bill|No DN|||
SOL1194501|SHP27-00615|To Deliver and Bill|No DN|||
SOL1194500|SHP27-00614|To Deliver and Bill|No DN|||
SOL1194498|SHP27-00611|To Deliver and Bill|No DN|||
SOL1194496|SHP27-00609|To Deliver and Bill|No DN|||
SOL1194495|SHP27-00608|To Deliver and Bill|No DN|||
SOL1194493|SHP27-00606|To Deliver and Bill|No DN|||
SOL1194491|SHP27-00604|To Deliver and Bill|No DN|||
SOL1194490|SHP27-00603|To Deliver and Bill|No DN|||
SOL1194489|SHP27-00602|To Deliver and Bill|No DN|||
SOL1194488|SHP27-00601|To Deliver and Bill|No DN|||
SOL1194487|SHP27-00600|To Deliver and Bill|No DN|||
SOL1194483|SHP27-00596|To Deliver and Bill|No DN|||
SOL1194482|SHP27-00595|To Deliver and Bill|No DN|||
SOL1194477|SHP27-00590|To Deliver and Bill|No DN|||
SOL1194476|SHP27-00589|To Deliver and Bill|No DN|||
SOL1194475|SHP27-00588|To Deliver and Bill|No DN|||
SOL1194474|SHP27-00587|To Deliver and Bill|No DN|||
SOL1194473|SHP27-00586|To Deliver and Bill|No DN|||
SOL1194472|SHP27-00585|To Deliver and Bill|No DN|||
SOL1194471|SHP27-00584|To Deliver and Bill|No DN|||
SOL1194470|SHP27-00583|To Deliver and Bill|No DN|||
SOL1194469|SHP27-00582|To Deliver and Bill|No DN|||
SOL1194467|SHP27-00580|To Deliver and Bill|No DN|||
SOL1194466|SHP27-00579|To Deliver and Bill|No DN|||
SOL1194465|SHP27-00578|To Deliver and Bill|No DN|||
SOL1194464|SHP27-00577|To Deliver and Bill|No DN|||
SOL1194463|SHP27-00576|To Deliver and Bill|No DN|||
SOL1194461|SHP27-00574|To Deliver and Bill|No DN|||
SOL1194460|SHP27-00573|To Deliver and Bill|No DN|||
SOL1194457|SHP27-00570|To Deliver and Bill|No DN|||
SOL1194456|SHP27-00569|To Deliver and Bill|No DN|||
SOL1194455|SHP27-00568|To Deliver and Bill|No DN|||
SOL1194454|SHP27-00567|To Deliver and Bill|No DN|||
SOL1194453|SHP27-00566|To Deliver and Bill|No DN|||
SOL1194451|SHP27-00564|To Deliver and Bill|No DN|||
SOL1194450|SHP27-00563|To Deliver and Bill|No DN|||
SOL1194449|SHP27-00562|To Deliver and Bill|No DN|||
SOL1194447|SHP27-00560|To Deliver and Bill|No DN|||
SOL1194446|SHP27-00559|To Deliver and Bill|No DN|||
SOL1194445|SHP27-00558|To Deliver and Bill|No DN|||
SOL1194443|SHP27-00556|To Deliver and Bill|No DN|||
SOL1194442|SHP27-00555|To Deliver and Bill|No DN|||
SOL1194440|SHP27-00553|To Deliver and Bill|No DN|||
SOL1194439|SHP27-00552|To Deliver and Bill|No DN|||
SOL1194437|SHP27-00550|To Deliver and Bill|No DN|||
SOL1194436|SHP27-00549|To Deliver and Bill|No DN|||
SOL1194434|SHP27-00547|To Deliver and Bill|No DN|||
SOL1194433|SHP27-00546|To Deliver and Bill|No DN|||
SOL1194431|SHP27-00544|To Deliver and Bill|No DN|||
SOL1194429|SHP27-00542|To Deliver and Bill|No DN|||
SOL1194426|SHP27-00539|To Deliver and Bill|No DN|||
SOL1194424|SHP27-00537|To Deliver and Bill|No DN|||
SOL1194423|SHP27-00536|To Deliver and Bill|No DN|||
SOL1194420|SHP27-00533|To Deliver and Bill|No DN|||
SOL1194419|SHP27-00532|To Deliver and Bill|No DN|||
SOL1194418|SHP27-00531|To Deliver and Bill|No DN|||
SOL1194415|SHP27-00528|To Deliver and Bill|No DN|||
SOL1194413|SHP27-00526|To Deliver and Bill|No DN|||
SOL1194412|SHP27-00525|To Deliver and Bill|No DN|||
SOL1194411|SHP27-00521|To Deliver and Bill|No DN|||
SOL1194410|SHP27-00520|To Deliver and Bill|No DN|||
SOL1194407|SHP27-00517|To Deliver and Bill|No DN|||
SOL1194406|SHP27-00516|To Deliver and Bill|No DN|||
SOL1194404|SHP27-00514|To Deliver and Bill|No DN|||
SOL1194401|SHP27-00512|To Deliver and Bill|No DN|||
SOL1194399|SHP27-00510|To Deliver and Bill|No DN|||
SOL1194397|SHP27-00508|To Deliver and Bill|No DN|||
SOL1194396|SHP27-00507|To Deliver and Bill|No DN|||
SOL1194395|SHP27-00506|To Deliver and Bill|No DN|||
SOL1194393|SHP27-00504|To Deliver and Bill|No DN|||
SOL1194392|SHP27-00503|To Deliver and Bill|No DN|||
SOL1194390|SHP27-00501|To Deliver and Bill|No DN|||
SOL1194388|SHP27-00499|To Deliver and Bill|No DN|||
SOL1194385|SHP27-00496|To Deliver and Bill|No DN|||
SOL1194384|SHP27-00495|To Deliver and Bill|No DN|||
SOL1194382|SHP27-00493|To Deliver and Bill|No DN|||
SOL1194381|SHP27-00492|To Deliver and Bill|No DN|||
SOL1194379|SHP27-00490|To Deliver and Bill|No DN|||
SOL1194378|SHP27-00489|To Deliver and Bill|No DN|||
SOL1194376|SHP27-00487|To Deliver and Bill|No DN|||
SOL1194375|SHP27-00486|To Deliver and Bill|No DN|||
SOL1194374|SHP27-00485|To Deliver and Bill|No DN|||
SOL1194373|SHP27-00484|To Deliver and Bill|No DN|||
SOL1194372|SHP27-00483|To Deliver and Bill|No DN|||
SOL1194369|SHP27-00480|To Deliver and Bill|No DN|||
SOL1194367|SHP27-00478|To Deliver and Bill|No DN|||
SOL1194366|SHP27-00477|To Deliver and Bill|No DN|||
SOL1194365|SHP27-00476|To Deliver and Bill|No DN|||
SOL1194364|SHP27-00475|To Deliver and Bill|No DN|||
SOL1194363|SHP27-00474|To Deliver and Bill|No DN|||
SOL1194360|SHP27-00471|To Deliver and Bill|No DN|||
SOL1194359|SHP27-00470|To Deliver and Bill|No DN|||
SOL1194358|NOT FOUND|||||
SOL1194356|SHP27-00468|To Deliver and Bill|No DN|||
SOL1194355|SHP27-00467|To Deliver and Bill|No DN|||
SOL1194353|SHP27-00465|To Deliver and Bill|No DN|||
SOL1194350|SHP27-00462|To Deliver and Bill|No DN|||
SOL1194346|SHP27-00458|To Deliver and Bill|No DN|||
SOL1194345|SHP27-00457|To Deliver and Bill|No DN|||
SOL1194344|NOT FOUND|||||
SOL1194343|SHP27-00455|To Deliver and Bill|No DN|||
SOL1194342|SHP27-00454|To Deliver and Bill|No DN|||
SOL1194341|SHP27-00453|To Deliver and Bill|No DN|||
SOL1194340|SHP27-00452|To Deliver and Bill|No DN|||
SOL1194338|SHP27-00450|To Deliver and Bill|No DN|||
SOL1194336|SHP27-00448|To Deliver and Bill|No DN|||
SOL1194335|SHP27-00447|To Deliver and Bill|No DN|||
SOL1194334|SHP27-00446|To Deliver and Bill|No DN|||
SOL1194333|SHP27-00445|To Deliver and Bill|No DN|||
SOL1194331|NOT FOUND|||||
SOL1194330|SHP27-00442|To Deliver and Bill|No DN|||
SOL1194329|SHP27-00441|To Deliver and Bill|No DN|||
SOL1194328|SHP27-00440|To Deliver and Bill|No DN|||
SOL1194326|SHP27-00438|To Deliver and Bill|No DN|||
SOL1194325|SHP27-00437|To Deliver and Bill|No DN|||
SOL1194324|SHP27-00436|To Deliver and Bill|No DN|||
SOL1194323|SHP27-00435|To Deliver and Bill|No DN|||
SOL1194321|SHP27-00433|To Deliver and Bill|No DN|||
SOL1194320|SHP27-00432|To Deliver and Bill|No DN|||
SOL1194319|SHP27-00431|To Deliver and Bill|No DN|||
SOL1194318|NOT FOUND|||||
SOL1194317|SHP27-00429|To Deliver and Bill|No DN|||
SOL1194314|SHP27-00426|To Deliver and Bill|No DN|||
SOL1194313|SHP27-00425|To Deliver and Bill|No DN|||
SOL1194312|SHP27-00424|To Deliver and Bill|No DN|||
SOL1194310|NOT FOUND|||||
SOL1194309|SHP27-00421|To Deliver and Bill|No DN|||
SOL1194308|SHP27-00420|To Deliver and Bill|No DN|||
SOL1194307|SHP27-00419|To Deliver and Bill|No DN|||
SOL1194306|SHP27-00418|To Deliver and Bill|No DN|||
SOL1194305|SHP27-00417|To Deliver and Bill|No DN|||
SOL1194304|SHP27-00416|To Deliver and Bill|No DN|||
SOL1194303|SHP27-00415|To Deliver and Bill|No DN|||
SOL1194302|SHP27-00414|To Deliver and Bill|No DN|||
SOL1194299|NOT FOUND|||||
SOL1194298|SHP27-00410|To Deliver and Bill|No DN|||
SOL1194296|SHP27-00408|To Deliver and Bill|No DN|||
SOL1194295|SHP27-00407|To Deliver and Bill|No DN|||
SOL1194293|SHP27-00405|To Deliver and Bill|No DN|||
SOL1194292|NOT FOUND|||||
SOL1194290|SHP27-00402|To Deliver and Bill|No DN|||
SOL1194289|SHP27-00401|To Deliver and Bill|No DN|||
SOL1194288|SHP27-00400|To Deliver and Bill|No DN|||
SOL1194287|SHP27-00399|To Deliver and Bill|No DN|||
SOL1194286|SHP27-00398|To Deliver and Bill|No DN|||
SOL1194285|SHP27-00397|To Deliver and Bill|No DN|||
SOL1194284|SHP27-00396|To Deliver and Bill|No DN|||
SOL1194281|SHP27-00395-1|Draft|No DN|||
SOL1194280|SHP27-00395|To Deliver and Bill|No DN|||
SOL1194279|SHP27-00395|To Deliver and Bill|No DN|||
SOL1194278|SHP27-00395|To Deliver and Bill|No DN|||
SOL1194277|NOT FOUND|||||
SOL1194276|SHP27-00395|To Deliver and Bill|No DN|||
SOL1194275|SHP27-00395|To Deliver and Bill|No DN|||
SOL1194274|NOT FOUND|||||
SOL1194272|SHP27-00395|To Deliver and Bill|No DN|||
SOL1194271|SHP27-00395|To Deliver and Bill|No DN|||
SOL1194269|SHP27-00395|To Deliver and Bill|No DN|||
SOL1194268|SHP27-00394-1|Draft|No DN|||
SOL1194267|SHP27-00393|To Deliver and Bill|No DN|||
SOL1194266|SHP27-00392|To Deliver and Bill|No DN|||
SOL1194265|SHP27-00391|To Deliver and Bill|No DN|||
SOL1194264|SHP27-00390|To Deliver and Bill|No DN|||
SOL1194263|SHP27-00389|To Deliver and Bill|No DN|||
SOL1194262|SHP27-00388|To Deliver and Bill|No DN|||
SOL1194261|SHP27-00387|To Deliver and Bill|No DN|||
SOL1194258|SHP27-00384|To Deliver and Bill|No DN|||
SOL1194257|SHP27-00383|To Deliver and Bill|No DN|||
SOL1194256|SHP27-00382|To Deliver and Bill|No DN|||
SOL1194255|SHP27-00381|To Deliver and Bill|No DN|||
SOL1194254|SHP27-00380|To Deliver and Bill|No DN|||
SOL1194253|SHP27-00379|To Deliver and Bill|No DN|||
SOL1194250|SHP27-00376-1|Draft|No DN|||
SOL1194249|SHP27-00375|To Deliver and Bill|No DN|||
SOL1194248|SHP27-00374|To Deliver and Bill|No DN|||
SOL1194247|SHP27-00373|To Deliver and Bill|No DN|||
SOL1194245|SHP27-00371|To Deliver and Bill|No DN|||
SOL1194244|SHP27-00370|To Deliver and Bill|No DN|||
SOL1194243|SHP27-00369|To Deliver and Bill|No DN|||
SOL1194240|SHP27-00366|To Deliver and Bill|No DN|||
SOL1194239|SHP27-00365|To Deliver and Bill|No DN|||
SOL1194237|SHP27-00363|To Deliver and Bill|No DN|||
SOL1194236|SHP27-00362|To Deliver and Bill|No DN|||
SOL1194235|SHP27-00361|To Deliver and Bill|No DN|||
SOL1194234|SHP27-00360|To Deliver and Bill|No DN|||
SOL1194233|SHP27-00359|To Deliver and Bill|No DN|||
SOL1194231|SHP27-00357|To Deliver and Bill|No DN|||
SOL1194229|SHP27-00355|To Deliver and Bill|No DN|||
SOL1194228|SHP27-00354|To Deliver and Bill|No DN|||
SOL1194226|SHP27-00352|To Deliver and Bill|No DN|||
SOL1194225|SHP27-00351|To Deliver and Bill|No DN|||
SOL1194224|SHP27-00350|To Deliver and Bill|No DN|||
SOL1194221|SHP27-00347|To Deliver and Bill|No DN|||
SOL1194218|SHP27-00344|To Deliver and Bill|No DN|||
SOL1194217|SHP27-00343|To Deliver and Bill|No DN|||
SOL1194213|SHP27-00339|To Deliver and Bill|No DN|||
SOL1194212|SHP27-00338|To Deliver and Bill|No DN|||
SOL1194211|SHP27-00337|To Deliver and Bill|No DN|||
SOL1194210|SHP27-00336|To Deliver and Bill|No DN|||
SOL1194206|SHP27-00332|To Deliver and Bill|No DN|||
SOL1194204|SHP27-00330|To Deliver and Bill|No DN|||
SOL1194202|SHP27-00328|To Deliver and Bill|No DN|||
SOL1194201|SHP27-00327|To Deliver and Bill|No DN|||
SOL1194197|SHP27-00323|To Deliver and Bill|No DN|||
SOL1194196|SHP27-00322|To Deliver and Bill|No DN|||
SOL1194193|SHP27-00319|To Deliver and Bill|No DN|||
SOL1194192|SHP27-00318|To Deliver and Bill|No DN|||
SOL1194190|SHP27-00316|To Deliver and Bill|No DN|||
SOL1194189|SHP27-00315|To Deliver and Bill|No DN|||
SOL1194188|SHP27-00314|To Deliver and Bill|No DN|||
SOL1194187|SHP27-00313|To Deliver and Bill|No DN|||
SOL1194186|SHP27-00312|To Deliver and Bill|No DN|||
SOL1194184|SHP27-00310|To Deliver and Bill|No DN|||
SOL1194183|SHP27-00309|To Deliver and Bill|No DN|||
SOL1194182|SHP27-00308|To Deliver and Bill|No DN|||
SOL1194181|SHP27-00307|To Deliver and Bill|No DN|||
SOL1194180|SHP27-00306|To Deliver and Bill|No DN|||
SOL1194179|SHP27-00305|To Deliver and Bill|No DN|||
SOL1194178|SHP27-00304|To Deliver and Bill|No DN|||
SOL1194177|SHP27-00303|To Deliver and Bill|No DN|||
SOL1194176|SHP27-00302|To Deliver and Bill|No DN|||
SOL1194175|SHP27-00301|To Deliver and Bill|No DN|||
SOL1194174|SHP27-00300|To Deliver and Bill|No DN|||
SOL1194173|SHP27-00299|To Deliver and Bill|No DN|||
SOL1194172|SHP27-00298|To Deliver and Bill|No DN|||
SOL1194168|SHP27-00294|To Deliver and Bill|No DN|||
SOL1194167|SHP27-00293|To Deliver and Bill|No DN|||
SOL1194166|SHP27-00292|To Deliver and Bill|No DN|||
SOL1194164|SHP27-00290|To Deliver and Bill|No DN|||
SOL1194163|SHP27-00289|To Deliver and Bill|No DN|||
SOL1194161|SHP27-00287|To Deliver and Bill|No DN|||
SOL1194160|SHP27-00286|To Deliver and Bill|No DN|||
SOL1194159|SHP27-00285|To Deliver and Bill|No DN|||
SOL1194158|SHP27-00284|To Deliver and Bill|No DN|||
SOL1194157|SHP27-00283|To Deliver and Bill|No DN|||
SOL1194155|SHP27-00281|To Deliver and Bill|No DN|||
SOL1194154|SHP27-00280|To Deliver and Bill|No DN|||
SOL1194152|SHP27-00278|To Deliver and Bill|No DN|||
SOL1194151|SHP27-00277|To Deliver and Bill|No DN|||
SOL1194150|SHP27-00276|To Deliver and Bill|No DN|||
SOL1194149|SHP27-00275|To Deliver and Bill|No DN|||
SOL1194148|SHP27-00274|To Deliver and Bill|No DN|||
SOL1194147|SHP27-00273|To Deliver and Bill|No DN|||
SOL1194146|SHP27-00272|To Deliver and Bill|No DN|||
SOL1194145|SHP27-00271|To Deliver and Bill|No DN|||
SOL1194144|SHP27-00270|To Deliver and Bill|No DN|||
SOL1194143|SHP27-00269|To Deliver and Bill|No DN|||
SOL1194142|SHP27-00268|To Deliver and Bill|No DN|||
SOL1194141|SHP27-00267|To Deliver and Bill|No DN|||
SOL1194140|SHP27-00266|To Deliver and Bill|No DN|||
SOL1194139|SHP27-00265|To Deliver and Bill|No DN|||
SOL1194138|SHP27-00264|To Deliver|No DN|||
SOL1194137|SHP27-00263|To Deliver|No DN|||
SOL1194136|SHP27-00262|To Deliver|No DN|||
SOL1194135|SHP27-00261|To Deliver|No DN|||
SOL1194134|SHP27-00260|To Deliver|No DN|||
SOL1194133|SHP27-00259|To Deliver|No DN|||
SOL1194131|SHP27-00257|To Deliver|No DN|||
SOL1194130|SHP27-00256|To Deliver|No DN|||
SOL1194128|SHP27-00254|To Deliver|No DN|||
SOL1194126|SHP27-00252-1|Draft|No DN|||
SOL1194124|SHP27-00250|To Deliver|No DN|||
SOL1194123|SHP27-00249|To Deliver|No DN|||
SOL1194122|SHP27-00248|To Deliver|No DN|||
SOL1194114|SHP27-00240-1|Draft|No DN|||
SOL1194113|SHP27-00239|To Deliver|SHPDN27-00463|Draft||
SOL1194106|SHP27-00232-1|Draft|No DN|||
SOL1194057|SHP27-00181-1|Draft|No DN|||
SOL1194052|SHP27-00176-1|Draft|No DN|||
SOL1194030|SHP27-00098|Cancelled|No DN|||
SOL1194019|SHP27-00087|Completed|SHPDN27-00231|Completed||
SOL1194017|SHP27-00085|Completed|SHPDN27-00229|Completed||
SOL1194016|SHP27-00084|Cancelled|No DN|||
SOL1194004|SHP27-00072|Completed|SHPDN27-00218|Completed||
SOL1193997|SHP27-00065|Completed|SHPDN27-00211|Completed||
SOL1193984|SHP27-00105|Completed|SHPDN27-00246|Completed||
SOL1193982|SHP27-00103|Completed|SHPDN27-00244|Completed||
SOL1193979|SHP27-00100|Completed|SHPDN27-00241|Completed||
SOL1193978|SHP27-00099|Completed|SHPDN27-00240|Completed||
SOL1193976|SHP27-00097|Completed|SHPDN27-00239|Completed||
SOL1193958|SHP27-00079|Completed|SHPDN27-00225|Completed||
SOL1193952|SHP27-00073|Completed|SHPDN27-00219|Completed||
SOL1193941|SHP27-00062|To Deliver|SHPDN27-00208|Draft||
SOL1193935|SHP27-00056|Completed|SHPDN27-00202|Completed||
SOL1193932|SHP27-00053|Completed|SHPDN27-00199|Completed||
SOL1193914|SHP27-00035|Completed|SHPDN27-00181|Completed||
SOL1193902|SHP27-00023|Completed|SHPDN27-00169|Completed||
SOL1193894|SHP27-00015|Completed|SHPDN27-00161|Completed||
SOL1193891|SHP27-00012|Completed|SHPDN27-00158|Completed||
SOL1193889|SHP27-00010|Completed|SHPDN27-00156|Completed||
SOL1193874|SHP-2026-2027-00265|Completed|SHPDN-2026-2027-00255|Completed||
SOL1193873|SHP-2026-2027-00264|Completed|SHPDN-2026-2027-00254|Completed||
SOL1193868|SHP-2026-2027-00259|Completed|SHPDN-2026-2027-00249|Completed||
SOL1193865|SHP-2026-2027-00256|Completed|SHPDN-2026-2027-00246|Completed||
SOL1193842|SHP-2026-2027-00237|To Deliver|SHPDN-2026-2027-00220|Draft||
SOL1193839|SHP-2026-2027-00234|Completed|SHPDN-2026-2027-00217|Completed||
SOL1193820|SHP-2026-2027-00218|Completed|SHPDN-2026-2027-00202|Completed||
SOL1193805|SHP-2026-2027-00204|Completed|SHPDN-2026-2027-00192|Completed||
SOL1193795|SHP-2026-2027-00195|Cancelled|No DN|||
SOL1193787|SHP-2026-2027-00191|Cancelled|SHPDN-2026-2027-00179|Cancelled||
SOL1193773|SHP-2026-2027-00178|Cancelled|SHPDN-2026-2027-00168|Cancelled||
SOL1193772|SHP-2026-2027-00177|Cancelled|SHPDN-2026-2027-00167|Cancelled||
SOL1193754|SHP-2026-2027-00162|Completed|SHPDN-2026-2027-00150|Completed||
SOL1193753|SHP-2026-2027-00161|Completed|SHPDN-2026-2027-00149|Completed||
SOL1193748|SHP-2026-2027-00158|Completed|SHPDN27-00415|Completed||
SOL1193745|SHP-2026-2027-00156|Completed|SHPDN-2026-2027-00144|Completed||
SOL1193732|SHP-2026-2027-00145|Cancelled|SHPDN-2026-2027-00135|Cancelled||
SOL1193726|SHP-2026-2027-00140|Cancelled|SHPDN-2026-2027-00131|Cancelled||
SOL1193720|SHP-2026-2027-00135|Completed|SHPDN-2026-2027-00126|Completed||
SOL1193719|SHP-2026-2027-00134|Completed|SHPDN-2026-2027-00125|Completed||
SOL1193716|SHP27-00139|Completed|SHPDN27-00278|Completed||
SOL1193684|SHP27-00141|To Deliver|SHPDN27-00280|Draft||
SOL1193683|SHP-2026-2027-00186|To Deliver|SHPDN27-00063|Draft||
SOL1193679|SHP-2026-2027-00183|Cancelled|No DN|||
SOL1193678|SHP-2026-2027-00182|Cancelled|No DN|||
SOL1193677|SHP27-00143|Completed|SHPDN27-00282|Completed||
SOL1193674|SHP27-00144|Completed|SHPDN27-00283|Completed||
SOL1193661|SHP27-00146|Completed|SHPDN27-00285|Completed||
SOL1193658|SHP27-00147|Completed|SHPDN27-00286|Completed||
SOL1193645|SHP27-00149|Completed|SHPDN27-00288|Completed||
SOL1193642|SHP27-00150|Completed|SHPDN27-00289|Completed||
SOL1193641|SHP27-00151|Completed|SHPDN27-00290|Completed||
SOL1193639|SHP-2026-2027-00153|Cancelled|SHPDN-2026-2027-00141|Cancelled||
SOL1193623|SHP-2026-2027-00141|Cancelled|SHPDN-2026-2027-00132|Cancelled||
SOL1193617|SHP27-00153|Completed|SHPDN27-00291|Completed||
SOL1193587|SHP-2026-2027-00119|Cancelled|No DN|||
SOL1193581|SHP-2026-2027-00113|Cancelled|SHPDN-2026-2027-00104|Cancelled||
SOL1193572|SHP-2026-2027-00108|Cancelled|SHPDN-2026-2027-00099|Cancelled||
SOL1193560|SHP-2026-2027-00099|Completed|SHPDN-2026-2027-00090|Completed||
SOL1193556|SHP-2026-2027-00096|Completed|SHPDN-2026-2027-00087|Completed||
SOL1193542|SHP-2026-2027-00085|Completed|SHPDN-2026-2027-00079|Completed||
SOL1193532|SHP-2026-2027-00078|Completed|SHPDN-2026-2027-00074|Completed||
SOL1193528|SHP-2026-2027-00076|Completed|SHPDN-2026-2027-00072|Completed||
SOL1193523|SHP-2026-2027-00072|Cancelled|SHPDN-2026-2027-00068|Cancelled||
SOL1193516|SHP-2026-2027-00066|Completed|SHPDN-2026-2027-00063|Completed||
SOL1193514|SHP-2026-2027-00064|Completed|SHPDN-2026-2027-00061|Completed||
SOL1193513|SHP-2026-2027-00063|Completed|SHPDN-2026-2027-00060|Completed||
SOL1193505|SHP-2026-2027-00058|Completed|SHPDN-2026-2027-00055|Completed||
SOL1193500|SHP-2026-2027-00055|Completed|SHPDN-2026-2027-00052|Completed||
SOL1193498|SHP-2026-2027-00054|Completed|SHPDN-2026-2027-00051|Completed||"""

# De-duplicate rows (some SOs have multiple DN items pointing to same DN)
seen = set()
rows = []
for line in RAW.strip().split('\n'):
    parts = line.split('|')
    key = f"{parts[0]}|{parts[3]}"  # order_id + DN
    if key not in seen:
        seen.add(key)
        rows.append(parts)

# Build Excel
wb = openpyxl.Workbook()

# ── Main Sheet ──
ws = wb.active
ws.title = "Shopify Orders"

headers = ["Order ID", "SO Number", "SO Status", "DN Number", "DN Status", "AWB", "Tracking Status"]
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center")
thin_border = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

# Status colors
status_fills = {
    "Completed": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "To Deliver and Bill": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "To Deliver": PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid"),
    "Draft": PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid"),
    "Cancelled": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "NOT FOUND": PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid"),
    "No DN": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
}
status_fonts = {
    "Completed": Font(color="006100"),
    "Cancelled": Font(color="9C0006"),
    "NOT FOUND": Font(color="990000", bold=True),
}

for r_idx, parts in enumerate(rows, 2):
    for c_idx, val in enumerate(parts, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=val.strip())
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # Color SO Status column
        if c_idx == 3:
            fill = status_fills.get(val.strip())
            if fill:
                cell.fill = fill
            font = status_fonts.get(val.strip())
            if font:
                cell.font = font
        # Color DN Status column
        if c_idx == 5:
            fill = status_fills.get(val.strip())
            if fill:
                cell.fill = fill
            font = status_fonts.get(val.strip())
            if font:
                cell.font = font
        # Color DN Number "No DN"
        if c_idx == 4 and val.strip() == "No DN":
            cell.fill = status_fills["No DN"]
            cell.font = Font(color="808080", italic=True)
        # Color Order ID "NOT FOUND"
        if c_idx == 2 and val.strip() == "NOT FOUND":
            cell.fill = status_fills["NOT FOUND"]
            cell.font = status_fonts["NOT FOUND"]

# Column widths
col_widths = [16, 28, 22, 30, 16, 20, 18]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Freeze top row + auto filter
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:G{len(rows) + 1}"

# ── Summary Sheet ──
ws2 = wb.create_sheet("Summary")
ws2.sheet_properties.tabColor = "2F5496"

# Count statuses
so_status_counts = Counter()
dn_status_counts = Counter()
not_found = 0
no_dn = 0
has_awb = 0

for parts in rows:
    so_status = parts[2].strip()
    dn_val = parts[3].strip()
    dn_status = parts[4].strip() if len(parts) > 4 else ""
    awb = parts[5].strip() if len(parts) > 5 else ""

    if so_status == "":
        so_status_counts["NOT FOUND"] += 1
        not_found += 1
    else:
        so_status_counts[so_status] += 1

    if dn_val == "No DN" or dn_val == "":
        no_dn += 1
    else:
        dn_status_counts[dn_status if dn_status else "Unknown"] += 1

    if awb:
        has_awb += 1

summary_data = [
    ["Metric", "Count"],
    ["Total Orders (input)", 380],
    ["Total Rows (with splits)", len(rows)],
    ["NOT FOUND in Atlas", not_found],
    ["", ""],
    ["SO Status Breakdown", ""],
]
for status, count in sorted(so_status_counts.items(), key=lambda x: -x[1]):
    summary_data.append([f"  {status}", count])

summary_data.extend([
    ["", ""],
    ["DN Status Breakdown", ""],
])
for status, count in sorted(dn_status_counts.items(), key=lambda x: -x[1]):
    summary_data.append([f"  {status}", count])

summary_data.extend([
    ["  No DN", no_dn],
    ["", ""],
    ["Orders with AWB", has_awb],
])

for r_idx, (label, val) in enumerate(summary_data, 1):
    c1 = ws2.cell(row=r_idx, column=1, value=label)
    c2 = ws2.cell(row=r_idx, column=2, value=val)
    c1.border = thin_border
    c2.border = thin_border
    c2.alignment = Alignment(horizontal="center")
    if r_idx == 1 or label in ["SO Status Breakdown", "DN Status Breakdown"]:
        c1.font = Font(bold=True, size=11)
        if r_idx == 1:
            c1.fill = header_fill
            c1.font = header_font
            c2.fill = header_fill
            c2.font = header_font

ws2.column_dimensions["A"].width = 30
ws2.column_dimensions["B"].width = 12

output_path = r"C:\Users\accou\Downloads\Shopify_Orders_Atlas_Report.xlsx"
wb.save(output_path)
print(f"Saved to {output_path}")
print(f"Total rows: {len(rows)}")
print(f"SO statuses: {dict(so_status_counts)}")
print(f"DN statuses: {dict(dn_status_counts)}")
print(f"No DN: {no_dn}, NOT FOUND: {not_found}, Has AWB: {has_awb}")

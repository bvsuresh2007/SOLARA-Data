import requests, json, time, re
from dotenv import dotenv_values
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

env = dotenv_values('C:/Users/accou/Documents/Projects/SOLARA-Data/.env')
s = requests.Session()
s.headers.update({
    'Authorization': f'token {env["ERPNEXT_API_KEY"]}:{env["ERPNEXT_API_SECRET"]}',
    'Content-Type': 'application/json'
})
BASE = env['ERPNEXT_URL']

# Original 51 DNs
original_51 = [
    "SHPDN27-00029","SHPDN27-00031","SHPDN27-00035","SHPDN27-00043","SHPDN27-00045",
    "SHPDN27-00049","SHPDN27-00057","SHPDN27-00059","SHPDN27-00078","SHPDN27-00087",
    "SHPDN27-00089","SHPDN27-00091","SHPDN27-00093","SHPDN27-00095","SHPDN27-00097",
    "SHPDN27-00099","SHPDN27-00101","SHPDN27-00103","SHPDN27-00105","SHPDN27-00107",
    "SHPDN27-00109","SHPDN27-00111","SHPDN27-00113","SHPDN27-00135","SHPDN27-00151",
    "SHPDN27-00152","SHPDN27-00158","SHPDN27-00176","SHPDN27-00181","SHPDN27-00198",
    "SHPDN27-00211","SHPDN27-00218","SHPDN27-00229","SHPDN27-00231","SHPDN27-00246",
    "SHPDN27-00263","SHPDN27-00264","SHPDN27-00265","SHPDN27-00270","SHPDN27-00273",
    "SHPDN27-00275","SHPDN27-00277","SHPDN27-00278","SHPDN27-00282","SHPDN27-00283",
    "SHPDN27-00285","SHPDN27-00286","SHPDN27-00474","SHPDN27-00488",
    "SHPDN-2026-2027-00072","SHPDN-2026-2027-00217"
]

# DNs we amended
amended_dns = set([
    "SHPDN27-00029","SHPDN27-00031","SHPDN27-00035","SHPDN27-00043","SHPDN27-00045",
    "SHPDN27-00049","SHPDN27-00057","SHPDN27-00059","SHPDN27-00078",
    "SHPDN27-00135","SHPDN27-00151","SHPDN27-00152","SHPDN27-00158","SHPDN27-00176",
    "SHPDN27-00181","SHPDN27-00198","SHPDN27-00211","SHPDN27-00218","SHPDN27-00229",
    "SHPDN27-00231","SHPDN27-00246","SHPDN27-00263","SHPDN27-00264","SHPDN27-00265",
    "SHPDN27-00270","SHPDN27-00273","SHPDN27-00275","SHPDN27-00277","SHPDN27-00278",
    "SHPDN27-00282","SHPDN27-00283","SHPDN27-00285","SHPDN27-00286","SHPDN27-00474",
    "SHPDN27-00488"
])

rows = []
for i, dn_name in enumerate(original_51):
    print(f'[{i+1}/{len(original_51)}] Fetching {dn_name}...', flush=True)

    # If amended, get the -1 version
    if dn_name in amended_dns:
        active_name = dn_name + "-1"
        was_amended = True
    else:
        active_name = dn_name
        was_amended = False

    r = s.get(f'{BASE}/api/resource/Delivery Note/{active_name}')
    if r.status_code != 200:
        r = s.get(f'{BASE}/api/resource/Delivery Note/{dn_name}')
        active_name = dn_name
        was_amended = False

    if r.status_code != 200:
        rows.append({
            'original_dn': dn_name, 'active_dn': '', 'amended': '',
            'status': 'NOT FOUND', 'docstatus': '', 'shipment_status': '',
            'awb': '', 'courier': '', 'tracking_url': '', 'shipping_label': '',
            'customer': '', 'shopify_order': '', 'so_name': '',
            'posting_date': '', 'items': '', 'total_qty': '', 'grand_total': '',
            'shipping_address': '', 'remarks': 'DN not found'
        })
        continue

    dn = r.json()['data']

    # Get SO from items
    so_name = ''
    item_names = []
    for item in dn.get('items', []):
        if item.get('against_sales_order') and not so_name:
            so_name = item['against_sales_order']
        item_names.append(f"{item.get('item_code','')} x {int(item.get('qty',0))}")

    # Clean address
    addr = dn.get('shipping_address', '')
    if addr:
        addr = re.sub(r'<br\s*/?>', ', ', addr)
        addr = re.sub(r'<[^>]+>', '', addr)
        addr = addr.strip().rstrip(',')

    # Determine remarks
    remarks = ''
    if was_amended:
        remarks = f'Amended from {dn_name} (was Failed)'
    if dn.get('docstatus') == 0:
        remarks = 'Draft - negative stock (SOL-JUC-121)'
    if active_name == 'SHPDN27-00488-1' and dn.get('shipment_status') == 'Failed':
        remarks = 'Clickpost rejection - 7.35kg rural Gujarat'

    rows.append({
        'original_dn': dn_name,
        'active_dn': active_name,
        'amended': 'Yes' if was_amended else 'No',
        'status': dn.get('status', ''),
        'docstatus': {0: 'Draft', 1: 'Submitted', 2: 'Cancelled'}.get(dn.get('docstatus'), ''),
        'shipment_status': dn.get('shipment_status', '') or '',
        'awb': dn.get('awb_number', '') or '',
        'courier': dn.get('courier_partner', '') or '',
        'tracking_url': dn.get('tracking_url', '') or '',
        'shipping_label': dn.get('shipping_label', '') or '',
        'customer': dn.get('customer_name', ''),
        'shopify_order': dn.get('custom_shopify_order_number', '') or '',
        'so_name': so_name,
        'posting_date': str(dn.get('posting_date', '')),
        'items': '; '.join(item_names),
        'total_qty': dn.get('total_qty', ''),
        'grand_total': dn.get('grand_total', ''),
        'shipping_address': addr,
        'remarks': remarks
    })
    time.sleep(0.3)

print(f'\nFetched {len(rows)} DNs. Building Excel...')

# Build Excel
wb = Workbook()
ws = wb.active
ws.title = "51 DN Status Report"

headers = [
    'S.No', 'Original DN', 'Active DN', 'Amended?', 'DN Status', 'Doc Status',
    'Shipment Status', 'AWB Number', 'Courier', 'Tracking URL', 'Shipping Label',
    'Customer', 'Shopify Order', 'Sales Order', 'Posting Date',
    'Items', 'Total Qty', 'Grand Total', 'Shipping Address', 'Remarks'
]

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

for i, row in enumerate(rows, 2):
    vals = [
        i - 1, row['original_dn'], row['active_dn'], row['amended'],
        row['status'], row['docstatus'], row['shipment_status'],
        row['awb'], row['courier'], row['tracking_url'], row['shipping_label'],
        row['customer'], row['shopify_order'], row['so_name'],
        row['posting_date'], row['items'], row['total_qty'],
        row['grand_total'], row['shipping_address'], row['remarks']
    ]
    for col, v in enumerate(vals, 1):
        cell = ws.cell(row=i, column=col, value=v)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=(col in [16, 19, 20]))

        # Color code shipment status
        if col == 7:
            if v == 'Created':
                cell.fill = green_fill
            elif v == 'Failed':
                cell.fill = red_fill
            elif not v:
                cell.fill = yellow_fill

        # Color code AWB
        if col == 8:
            if v:
                cell.fill = green_fill
            else:
                cell.fill = red_fill

        # Hyperlinks for tracking URL and shipping label
        if col in (10, 11) and v:
            cell.hyperlink = v
            cell.font = Font(color="0563C1", underline="single")

# Column widths
widths = {1:5, 2:20, 3:22, 4:10, 5:12, 6:12, 7:16, 8:20, 9:12, 10:45,
          11:45, 12:25, 13:16, 14:16, 15:12, 16:40, 17:9, 18:14, 19:50, 20:40}
for col, w in widths.items():
    ws.column_dimensions[get_column_letter(col)].width = w

ws.freeze_panes = 'A2'
ws.auto_filter.ref = f"A1:T{len(rows)+1}"
ws.row_dimensions[1].height = 30

# Summary sheet
ws2 = wb.create_sheet("Summary")
summary_data = [
    ('51 DN Status Report - Summary', ''),
    ('', ''),
    ('Total DNs', len(rows)),
    ('AWB Assigned', sum(1 for r in rows if r['awb'])),
    ('Shipment Created', sum(1 for r in rows if r['shipment_status'] == 'Created')),
    ('Shipment Failed', sum(1 for r in rows if r['shipment_status'] == 'Failed')),
    ('Draft (No Shipment)', sum(1 for r in rows if r['docstatus'] == 'Draft')),
    ('Amended DNs', sum(1 for r in rows if r['amended'] == 'Yes')),
    ('', ''),
    ('Generated', '2026-04-04'),
]
for i, (k, v) in enumerate(summary_data, 1):
    c1 = ws2.cell(row=i, column=1, value=k)
    c2 = ws2.cell(row=i, column=2, value=v)
    if i in (1, 3, 4, 5, 6, 7, 8):
        c1.font = Font(bold=True)
    if i == 1:
        c1.font = Font(bold=True, size=14)
ws2.column_dimensions['A'].width = 28
ws2.column_dimensions['B'].width = 15

output = 'C:/Users/accou/Downloads/51_DN_Status_Report.xlsx'
wb.save(output)
print(f'\nSaved to: {output}')

awb_count = sum(1 for r in rows if r['awb'])
created = sum(1 for r in rows if r['shipment_status'] == 'Created')
failed = sum(1 for r in rows if r['shipment_status'] == 'Failed')
draft = sum(1 for r in rows if r['docstatus'] == 'Draft')
print(f'AWB assigned: {awb_count}/51')
print(f'Created: {created} | Failed: {failed} | Draft: {draft}')

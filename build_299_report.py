import requests, json, time
from dotenv import dotenv_values
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

env = dotenv_values('C:/Users/accou/Documents/Projects/SOLARA-Data/.env')
s = requests.Session()
s.headers.update({
    'Authorization': f'token {env["ERPNEXT_API_KEY"]}:{env["ERPNEXT_API_SECRET"]}',
    'Content-Type': 'application/json'
})
BASE = env['ERPNEXT_URL']
TIMEOUT = 30

orders = """SOL1194557 SOL1194555 SOL1194554 SOL1194553 SOL1194551 SOL1194549 SOL1194548 SOL1194546 SOL1194544 SOL1194543
SOL1194542 SOL1194541 SOL1194539 SOL1194537 SOL1194536 SOL1194535 SOL1194534 SOL1194533 SOL1194530 SOL1194529
SOL1194528 SOL1194527 SOL1194525 SOL1194523 SOL1194522 SOL1194519 SOL1194517 SOL1194516 SOL1194515 SOL1194514
SOL1194513 SOL1194512 SOL1194511 SOL1194510 SOL1194509 SOL1194508 SOL1194507 SOL1194506 SOL1194505 SOL1194502
SOL1194501 SOL1194500 SOL1194498 SOL1194496 SOL1194495 SOL1194493 SOL1194491 SOL1194490 SOL1194489 SOL1194488
SOL1194487 SOL1194483 SOL1194482 SOL1194477 SOL1194476 SOL1194475 SOL1194474 SOL1194473 SOL1194472 SOL1194471
SOL1194470 SOL1194469 SOL1194467 SOL1194466 SOL1194465 SOL1194464 SOL1194463 SOL1194461 SOL1194460 SOL1194457
SOL1194456 SOL1194455 SOL1194454 SOL1194453 SOL1194451 SOL1194450 SOL1194449 SOL1194447 SOL1194446 SOL1194445
SOL1194443 SOL1194442 SOL1194440 SOL1194439 SOL1194437 SOL1194436 SOL1194434 SOL1194433 SOL1194431 SOL1194429
SOL1194426 SOL1194424 SOL1194423 SOL1194420 SOL1194419 SOL1194418 SOL1194415 SOL1194413 SOL1194412 SOL1194411
SOL1194410 SOL1194407 SOL1194406 SOL1194404 SOL1194401 SOL1194399 SOL1194397 SOL1194396 SOL1194395 SOL1194393
SOL1194392 SOL1194390 SOL1194388 SOL1194385 SOL1194384 SOL1194382 SOL1194381 SOL1194379 SOL1194378 SOL1194376
SOL1194375 SOL1194374 SOL1194373 SOL1194372 SOL1194369 SOL1194367 SOL1194366 SOL1194365 SOL1194364 SOL1194363
SOL1194360 SOL1194359 SOL1194356 SOL1194355 SOL1194353 SOL1194350 SOL1194346 SOL1194345 SOL1194343 SOL1194342
SOL1194341 SOL1194340 SOL1194338 SOL1194336 SOL1194335 SOL1194334 SOL1194333 SOL1194330 SOL1194329 SOL1194328
SOL1194326 SOL1194325 SOL1194324 SOL1194323 SOL1194321 SOL1194320 SOL1194319 SOL1194317 SOL1194314 SOL1194313
SOL1194312 SOL1194309 SOL1194308 SOL1194307 SOL1194306 SOL1194305 SOL1194304 SOL1194303 SOL1194302 SOL1194298
SOL1194296 SOL1194295 SOL1194293 SOL1194290 SOL1194289 SOL1194288 SOL1194287 SOL1194286 SOL1194285 SOL1194284
SOL1194281 SOL1194280 SOL1194279 SOL1194278 SOL1194276 SOL1194275 SOL1194272 SOL1194271 SOL1194269 SOL1194268
SOL1194267 SOL1194266 SOL1194265 SOL1194264 SOL1194263 SOL1194262 SOL1194261 SOL1194258 SOL1194257 SOL1194256
SOL1194255 SOL1194254 SOL1194253 SOL1194250 SOL1194249 SOL1194248 SOL1194247 SOL1194245 SOL1194244 SOL1194243
SOL1194240 SOL1194239 SOL1194237 SOL1194236 SOL1194235 SOL1194234 SOL1194233 SOL1194231 SOL1194229 SOL1194228
SOL1194226 SOL1194225 SOL1194224 SOL1194221 SOL1194218 SOL1194217 SOL1194213 SOL1194212 SOL1194211 SOL1194210
SOL1194206 SOL1194204 SOL1194202 SOL1194201 SOL1194197 SOL1194196 SOL1194193 SOL1194192 SOL1194190 SOL1194189
SOL1194188 SOL1194187 SOL1194186 SOL1194184 SOL1194183 SOL1194182 SOL1194181 SOL1194180 SOL1194179 SOL1194178
SOL1194177 SOL1194176 SOL1194175 SOL1194174 SOL1194173 SOL1194172 SOL1194168 SOL1194167 SOL1194166 SOL1194164
SOL1194163 SOL1194161 SOL1194160 SOL1194159 SOL1194158 SOL1194157 SOL1194155 SOL1194154 SOL1194152 SOL1194151
SOL1194150 SOL1194149 SOL1194148 SOL1194147 SOL1194146 SOL1194145 SOL1194144 SOL1194143 SOL1194142 SOL1194141
SOL1194140 SOL1194139 SOL1194138 SOL1194137 SOL1194136 SOL1194135 SOL1194134 SOL1194133 SOL1194131 SOL1194130
SOL1194128 SOL1194126 SOL1194124 SOL1194123 SOL1194122 SOL1194114 SOL1194106 SOL1194057 SOL1194052""".split()

print(f"Fetching data for {len(orders)} orders...")

# Batch lookup SOs
so_map = {}
batch_size = 20
for i in range(0, len(orders), batch_size):
    batch = orders[i:i+batch_size]
    r = s.get(f"{BASE}/api/resource/Sales Order", params={
        "filters": json.dumps([["custom_shopify_order_number","in",batch]]),
        "fields": json.dumps(["name","status","docstatus","custom_shopify_order_number","customer_name"]),
        "limit_page_length": batch_size
    }, timeout=TIMEOUT)
    for so in r.json().get("data", []):
        so_map[so["custom_shopify_order_number"]] = so
    time.sleep(0.2)
print(f"SOs found: {len(so_map)}")

# Lookup DNs
dn_map = {}
for so in so_map.values():
    r = s.get(f"{BASE}/api/resource/Delivery Note", params={
        "filters": json.dumps([["Delivery Note Item","against_sales_order","=",so["name"]],["docstatus","in",[0,1]]]),
        "fields": json.dumps(["name","status","docstatus","shipment_status","awb_number","courier_partner","tracking_url"]),
        "limit_page_length": 5
    }, timeout=TIMEOUT)
    dns = r.json().get("data", [])
    if dns:
        best = dns[0]
        for dn in dns:
            if dn.get("awb_number"):
                best = dn
                break
        dn_map[so["name"]] = best
    time.sleep(0.1)
print(f"DNs found: {len(dn_map)}")

# Build rows
rows = []
for oid in orders:
    so = so_map.get(oid)
    if not so:
        rows.append({"order_id": oid, "so_number": "", "so_status": "Not in Atlas",
                      "customer": "", "dn_number": "", "dn_status": "",
                      "shipment_status": "", "awb": "", "courier": "",
                      "tracking_url": "", "remarks": "Not synced from Shopify"})
        continue

    so_status_map = {0: "Draft", 2: "Cancelled"}
    so_display = so_status_map.get(so["docstatus"], so.get("status", ""))

    dn = dn_map.get(so["name"])
    if not dn:
        remarks = ""
        if so["docstatus"] == 0: remarks = "SO is Draft"
        elif so["docstatus"] == 2: remarks = "SO Cancelled"
        else: remarks = "No DN created"
        rows.append({"order_id": oid, "so_number": so["name"], "so_status": so_display,
                      "customer": so.get("customer_name", ""), "dn_number": "", "dn_status": "",
                      "shipment_status": "", "awb": "", "courier": "",
                      "tracking_url": "", "remarks": remarks})
        continue

    dn_status_map = {0: "Draft", 2: "Cancelled"}
    dn_display = dn_status_map.get(dn["docstatus"], dn.get("status", ""))
    awb = dn.get("awb_number", "") or ""
    ship = dn.get("shipment_status", "") or ""
    courier = dn.get("courier_partner", "") or ""
    tracking = dn.get("tracking_url", "") or ""

    remarks = ""
    if dn["docstatus"] == 0 and not awb:
        remarks = "DN Draft - needs submit"
    elif ship == "Failed":
        remarks = "Clickpost failed"

    rows.append({"order_id": oid, "so_number": so["name"], "so_status": so_display,
                  "customer": so.get("customer_name", ""), "dn_number": dn["name"],
                  "dn_status": dn_display, "shipment_status": ship, "awb": awb,
                  "courier": courier, "tracking_url": tracking, "remarks": remarks})

print(f"Building Excel with {len(rows)} rows...")

# Build Excel
wb = Workbook()
ws = wb.active
ws.title = "299 Orders Report"

headers = ["S.No", "Shopify Order ID", "SO Number", "SO Status", "Customer",
           "DN Number", "DN Status", "Shipment Status", "AWB Number",
           "Courier", "Tracking URL", "Remarks"]

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"), bottom=Side(style="thin"))

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border

for i, row in enumerate(rows, 2):
    vals = [i - 1, row["order_id"], row["so_number"], row["so_status"], row["customer"],
            row["dn_number"], row["dn_status"], row["shipment_status"], row["awb"],
            row["courier"], row["tracking_url"], row["remarks"]]
    for col, v in enumerate(vals, 1):
        cell = ws.cell(row=i, column=col, value=v)
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=(col == 12))

        if col == 9:  # AWB
            if v: cell.fill = green_fill
            elif row["dn_number"]: cell.fill = red_fill
            elif row["so_number"]: cell.fill = yellow_fill
            else: cell.fill = grey_fill

        if col == 8:  # Shipment
            if v == "Created": cell.fill = green_fill
            elif v == "Failed": cell.fill = red_fill

        if col == 4:  # SO Status
            if v == "Not in Atlas": cell.fill = grey_fill
            elif v == "Draft": cell.fill = yellow_fill
            elif v == "Cancelled": cell.fill = red_fill

        if col == 11 and v:  # Tracking URL
            cell.hyperlink = v
            cell.font = Font(color="0563C1", underline="single")

widths = {1:5, 2:16, 3:18, 4:22, 5:25, 6:20, 7:14, 8:16, 9:20, 10:12, 11:45, 12:30}
for col, w in widths.items():
    ws.column_dimensions[get_column_letter(col)].width = w

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:L{len(rows)+1}"
ws.row_dimensions[1].height = 30

# Summary sheet
ws2 = wb.create_sheet("Summary")
awb_count = sum(1 for r in rows if r["awb"])
no_awb_dn = sum(1 for r in rows if r["dn_number"] and not r["awb"])
no_dn = sum(1 for r in rows if r["so_number"] and not r["dn_number"] and r["so_status"] not in ("Draft","Cancelled","Not in Atlas"))
draft_so = sum(1 for r in rows if r["so_status"] == "Draft")
cancelled_so = sum(1 for r in rows if r["so_status"] == "Cancelled")
not_atlas = sum(1 for r in rows if r["so_status"] == "Not in Atlas")

summary = [
    ("299 Orders - AWB Status Report", ""),
    ("", ""),
    ("Total Orders", len(rows)),
    ("Has AWB", awb_count),
    ("DN exists, no AWB", no_awb_dn),
    ("No DN created", no_dn),
    ("Draft SO", draft_so),
    ("Cancelled SO", cancelled_so),
    ("Not in Atlas", not_atlas),
    ("", ""),
    ("Generated", "2026-04-04"),
]
for i, (k, v) in enumerate(summary, 1):
    c1 = ws2.cell(row=i, column=1, value=k)
    c2 = ws2.cell(row=i, column=2, value=v)
    if i == 1: c1.font = Font(bold=True, size=14)
    elif 3 <= i <= 9: c1.font = Font(bold=True)
ws2.column_dimensions["A"].width = 25
ws2.column_dimensions["B"].width = 15

output = "C:/Users/accou/Downloads/299_Orders_AWB_Report.xlsx"
wb.save(output)
print(f"\nSaved: {output}")
print(f"Summary: {awb_count} AWB | {no_awb_dn} DN no AWB | {no_dn} no DN | {draft_so} draft SO | {cancelled_so} cancelled | {not_atlas} not in Atlas")

import requests, json, time
from dotenv import dotenv_values
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

env = dotenv_values('C:/Users/accou/Documents/Projects/SOLARA-Data/.env')
s = requests.Session()
s.headers.update({
    'Authorization': f'token {env["ERPNEXT_API_KEY"]}:{env["ERPNEXT_API_SECRET"]}',
    'Content-Type': 'application/json'
})
BASE = env['ERPNEXT_URL']

orders = """SOL1194557 SOL1194555 SOL1194554 SOL1194553 SOL1194551 SOL1194549 SOL1194548 SOL1194546 SOL1194544 SOL1194543
SOL1194542 SOL1194541 SOL1194539 SOL1194537 SOL1194536 SOL1194535 SOL1194534 SOL1194533 SOL1194530 SOL1194529
SOL1194528 SOL1194527 SOL1194525 SOL1194523 SOL1194522 SOL1194519 SOL1194517 SOL1194516 SOL1194515 SOL1194514
SOL1194513 SOL1194512 SOL1194511 SOL1194510 SOL1194509 SOL1194508 SOL1194507 SOL1194506 SOL1194505 SOL1194502
SOL1194501 SOL1194500 SOL1194498 SOL1194496 SOL1194495 SOL1194493 SOL1194491 SOL1194490 SOL1194489 SOL1194488
SOL1194487 SOL1194483 SOL1194482 SOL1194477 SOL1194476 SOL1194475 SOL1194474 SOL1194473 SOL1194472 SOL1194471
SOL1194470 SOL1194469 SOL1194467 SOL1194466 SOL1194465 SOL1194464 SOL1194463 SOL1194461 SOL1194460 SOL1194457
SOL1194456 SOL1194455 SOL1194454 SOL1194453 SOL1194451 SOL1194450 SOL1194449 SOL1194447 SOL1194446 SOL1194445
SOL1194443 SOL1194442 SOL1194440 SOL1194439 SOL1194437 SOL1194436 SOL1194434 SOL1194433 SOL1194431 SOL1194429
SOL1194426 SOL1194424 SOL1194423 SOL1194420 SOL1194419 SOL1194418 SOL1194415 SOL1194413 SOL1194412 SOL1194411
SOL1194410 SOL1194407 SOL1194406 SOL1194404 SOL1194401 SOL1194399 SOL1194397 SOL1194396 SOL1194395 SOL1194393
SOL1194392 SOL1194390 SOL1194388 SOL1194385 SOL1194384 SOL1194382 SOL1194381 SOL1194379 SOL1194378 SOL1194376
SOL1194375 SOL1194374 SOL1194373 SOL1194372 SOL1194369 SOL1194367 SOL1194366 SOL1194365 SOL1194364 SOL1194363
SOL1194360 SOL1194359 SOL1194356 SOL1194355 SOL1194353 SOL1194350 SOL1194346 SOL1194345 SOL1194343 SOL1194342
SOL1194341 SOL1194340 SOL1194338 SOL1194336 SOL1194335 SOL1194334 SOL1194333 SOL1194330 SOL1194329 SOL1194328
SOL1194326 SOL1194325 SOL1194324 SOL1194323 SOL1194321 SOL1194320 SOL1194319 SOL1194317 SOL1194314 SOL1194313
SOL1194312 SOL1194309 SOL1194308 SOL1194307 SOL1194306 SOL1194305 SOL1194304 SOL1194303 SOL1194302 SOL1194298
SOL1194296 SOL1194295 SOL1194293 SOL1194290 SOL1194289 SOL1194288 SOL1194287 SOL1194286 SOL1194285 SOL1194284
SOL1194281 SOL1194280 SOL1194279 SOL1194278 SOL1194276 SOL1194275 SOL1194272 SOL1194271 SOL1194269 SOL1194268
SOL1194267 SOL1194266 SOL1194265 SOL1194264 SOL1194263 SOL1194262 SOL1194261 SOL1194258 SOL1194257 SOL1194256
SOL1194255 SOL1194254 SOL1194253 SOL1194250 SOL1194249 SOL1194248 SOL1194247 SOL1194245 SOL1194244 SOL1194243
SOL1194240 SOL1194239 SOL1194237 SOL1194236 SOL1194235 SOL1194234 SOL1194233 SOL1194231 SOL1194229 SOL1194228
SOL1194226 SOL1194225 SOL1194224 SOL1194221 SOL1194218 SOL1194217 SOL1194213 SOL1194212 SOL1194211 SOL1194210
SOL1194206 SOL1194204 SOL1194202 SOL1194201 SOL1194197 SOL1194196 SOL1194193 SOL1194192 SOL1194190 SOL1194189
SOL1194188 SOL1194187 SOL1194186 SOL1194184 SOL1194183 SOL1194182 SOL1194181 SOL1194180 SOL1194179 SOL1194178
SOL1194177 SOL1194176 SOL1194175 SOL1194174 SOL1194173 SOL1194172 SOL1194168 SOL1194167 SOL1194166 SOL1194164
SOL1194163 SOL1194161 SOL1194160 SOL1194159 SOL1194158 SOL1194157 SOL1194155 SOL1194154 SOL1194152 SOL1194151
SOL1194150 SOL1194149 SOL1194148 SOL1194147 SOL1194146 SOL1194145 SOL1194144 SOL1194143 SOL1194142 SOL1194141
SOL1194140 SOL1194139 SOL1194138 SOL1194137 SOL1194136 SOL1194135 SOL1194134 SOL1194133 SOL1194131 SOL1194130
SOL1194128 SOL1194126 SOL1194124 SOL1194123 SOL1194122 SOL1194114 SOL1194106 SOL1194057 SOL1194052""".split()

print(f"Fetching data for {len(orders)} orders...", flush=True)

# Batch lookup all SOs
all_sos = {}
batch_size = 20
for i in range(0, len(orders), batch_size):
    batch = orders[i:i+batch_size]
    r = s.get(f"{BASE}/api/resource/Sales Order", params={
        "filters": json.dumps([["custom_shopify_order_number","in",batch]]),
        "fields": json.dumps(["name","status","docstatus","custom_shopify_order_number"]),
        "limit_page_length": 100
    }, timeout=30)
    for so in r.json().get("data", []):
        oid = so["custom_shopify_order_number"]
        if oid not in all_sos:
            all_sos[oid] = []
        all_sos[oid].append(so)
    time.sleep(0.2)

# Pick best SO per order
so_map = {}
for oid, so_list in all_sos.items():
    best = None
    for so in so_list:
        if best is None:
            best = so
        elif so["docstatus"] > best["docstatus"]:
            best = so
        elif so["docstatus"] == best["docstatus"] and so["name"] > best["name"]:
            best = so
    so_map[oid] = best

print(f"SOs found: {len(so_map)}", flush=True)

# Lookup DNs
dn_map = {}
for oid, so in so_map.items():
    r = s.get(f"{BASE}/api/resource/Delivery Note", params={
        "filters": json.dumps([["Delivery Note Item","against_sales_order","=",so["name"]],["docstatus","in",[0,1]]]),
        "fields": json.dumps(["name","docstatus","awb_number"]),
        "limit_page_length": 5
    }, timeout=30)
    dns = r.json().get("data", [])
    if dns:
        best = dns[0]
        for dn in dns:
            if dn.get("awb_number"):
                best = dn
                break
        dn_map[so["name"]] = best
    time.sleep(0.1)

print(f"DNs found: {len(dn_map)}", flush=True)

# Build rows
rows = []
for oid in orders:
    so = so_map.get(oid)
    so_name = so["name"] if so else ""
    dn = dn_map.get(so_name) if so_name else None
    dn_name = dn["name"] if dn else ""
    awb = (dn.get("awb_number", "") or "") if dn else ""
    rows.append((oid, awb, dn_name, so_name))

# Build Excel
wb = Workbook()
ws = wb.active
ws.title = "299 Orders"

headers = ["Shopify Order ID", "AWB Number", "DN Number", "SO Number"]
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

for i, (oid, awb, dn, so) in enumerate(rows, 2):
    for col, v in enumerate([oid, awb, dn, so], 1):
        cell = ws.cell(row=i, column=col, value=v)
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center")

ws.column_dimensions["A"].width = 20
ws.column_dimensions["B"].width = 22
ws.column_dimensions["C"].width = 24
ws.column_dimensions["D"].width = 20
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:D{len(rows)+1}"

output = "C:/Users/accou/Downloads/299_Orders_AWB_List.xlsx"
wb.save(output)

awb_count = sum(1 for r in rows if r[1])
print(f"\nSaved: {output}")
print(f"Total: {len(rows)} | With AWB: {awb_count} | Without AWB: {len(rows)-awb_count}")

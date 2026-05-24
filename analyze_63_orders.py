import requests, json, time
from dotenv import dotenv_values
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

env = dotenv_values('C:/Users/accou/Documents/Projects/SOLARA-Data/.env')
s = requests.Session()
s.headers.update({
    'Authorization': f'token {env["ERPNEXT_API_KEY"]}:{env["ERPNEXT_API_SECRET"]}',
    'Content-Type': 'application/json'
})
BASE = env['ERPNEXT_URL']

orders = """SOL1194558 SOL1194556 SOL1194552 SOL1194550 SOL1194545 SOL1194540 SOL1194494 SOL1194481 SOL1194480 SOL1194459
SOL1194432 SOL1194430 SOL1194416 SOL1194409 SOL1194405 SOL1194394 SOL1194389 SOL1194386 SOL1194368 SOL1194361
SOL1194354 SOL1194352 SOL1194351 SOL1194348 SOL1194347 SOL1194327 SOL1194315 SOL1194311 SOL1194291 SOL1194246
SOL1194242 SOL1194222 SOL1194215 SOL1194214 SOL1194209 SOL1194208 SOL1194207 SOL1194205 SOL1194203 SOL1194199
SOL1194198 SOL1194195 SOL1194194 SOL1194185 SOL1194171 SOL1194170 SOL1194169 SOL1194162 SOL1194156 SOL1194127
SOL1194078 SOL1193959 SOL1193943 SOL1193919 SOL1193843 SOL1193766 SOL1193730 SOL1193702 SOL1193698 SOL1193692
SOL1193549 SOL1193536 SOL1193524""".split()

print(f"Fetching data for {len(orders)} orders...", flush=True)

# Batch lookup all SOs
all_sos = {}
batch_size = 20
for i in range(0, len(orders), batch_size):
    batch = orders[i:i+batch_size]
    r = s.get(f"{BASE}/api/resource/Sales Order", params={
        "filters": json.dumps([["custom_shopify_order_number","in",batch]]),
        "fields": json.dumps(["name","status","docstatus","custom_shopify_order_number"]),
        "limit_page_length": 100
    }, timeout=30)
    for so in r.json().get("data", []):
        oid = so["custom_shopify_order_number"]
        if oid not in all_sos:
            all_sos[oid] = []
        all_sos[oid].append(so)
    time.sleep(0.2)

# Pick best SO per order (submitted > draft, but NOT cancelled over submitted)
so_map = {}
for oid, so_list in all_sos.items():
    # Prefer submitted (1), then draft (0), then cancelled (2)
    priority = {1: 3, 0: 2, 2: 1}
    best = None
    for so in so_list:
        if best is None:
            best = so
        elif priority.get(so["docstatus"], 0) > priority.get(best["docstatus"], 0):
            best = so
        elif so["docstatus"] == best["docstatus"] and so["name"] > best["name"]:
            best = so
    so_map[oid] = best

print(f"SOs found: {len(so_map)}", flush=True)

# Lookup DNs for each SO
dn_map = {}
for oid, so in so_map.items():
    r = s.get(f"{BASE}/api/resource/Delivery Note", params={
        "filters": json.dumps([["Delivery Note Item","against_sales_order","=",so["name"]],["docstatus","in",[0,1]]]),
        "fields": json.dumps(["name","status","docstatus","shipment_status","awb_number","tracking_url"]),
        "limit_page_length": 5
    }, timeout=30)
    dns = r.json().get("data", [])
    if dns:
        best = dns[0]
        for dn in dns:
            if dn.get("awb_number"):
                best = dn
                break
        dn_map[so["name"]] = best
    time.sleep(0.1)

print(f"DNs found: {len(dn_map)}", flush=True)

# Build rows
rows = []
for oid in orders:
    so = so_map.get(oid)
    if not so:
        rows.append({"oid": oid, "so": "", "so_status": "Not in Atlas", "dn": "", "dn_status": "", "awb": "", "tracking": ""})
        continue

    so_status_map = {0: "Draft", 2: "Cancelled"}
    so_display = so_status_map.get(so["docstatus"], so.get("status", ""))

    dn = dn_map.get(so["name"])
    if not dn:
        rows.append({"oid": oid, "so": so["name"], "so_status": so_display, "dn": "", "dn_status": "", "awb": "", "tracking": ""})
        continue

    dn_status_map = {0: "Draft", 2: "Cancelled"}
    dn_display = dn_status_map.get(dn["docstatus"], dn.get("status", ""))
    awb = dn.get("awb_number", "") or ""
    ship = dn.get("shipment_status", "") or ""
    tracking = dn.get("tracking_url", "") or ""

    rows.append({
        "oid": oid,
        "so": so["name"],
        "so_status": so_display,
        "dn": dn["name"],
        "dn_status": dn_display,
        "awb": awb,
        "tracking": ship,
    })

print(f"Building Excel with {len(rows)} rows...", flush=True)

# Build Excel
wb = Workbook()
ws = wb.active
ws.title = "63 Orders Analysis"

headers = ["Order ID", "SO Number", "SO Status", "DN Number", "DN Status", "AWB", "Tracking Status"]
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

for i, row in enumerate(rows, 2):
    vals = [row["oid"], row["so"], row["so_status"], row["dn"], row["dn_status"], row["awb"], row["tracking"]]
    for col, v in enumerate(vals, 1):
        cell = ws.cell(row=i, column=col, value=v)
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center")

        # Color code AWB column
        if col == 6:
            if v:
                cell.fill = green_fill
            elif row["dn"]:
                cell.fill = red_fill
            elif row["so"]:
                cell.fill = yellow_fill
            else:
                cell.fill = grey_fill

        # Color code SO Status
        if col == 3:
            if v == "Not in Atlas":
                cell.fill = grey_fill
            elif v == "Draft":
                cell.fill = yellow_fill
            elif v == "Cancelled":
                cell.fill = red_fill

        # Color code Tracking Status
        if col == 7:
            if v == "Created":
                cell.fill = green_fill
            elif v == "Failed":
                cell.fill = red_fill

ws.column_dimensions["A"].width = 16
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 22
ws.column_dimensions["D"].width = 24
ws.column_dimensions["E"].width = 14
ws.column_dimensions["F"].width = 20
ws.column_dimensions["G"].width = 16
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:G{len(rows)+1}"

output = "C:/Users/accou/Downloads/63_Orders_Analysis.xlsx"
wb.save(output)

awb_count = sum(1 for r in rows if r["awb"])
no_so = sum(1 for r in rows if r["so_status"] == "Not in Atlas")
draft_so = sum(1 for r in rows if r["so_status"] == "Draft")
no_dn = sum(1 for r in rows if r["so"] and not r["dn"])
print(f"\nSaved: {output}")
print(f"Total: {len(rows)} | AWB: {awb_count} | No SO: {no_so} | Draft SO: {draft_so} | No DN: {no_dn} | DN no AWB: {sum(1 for r in rows if r['dn'] and not r['awb'])}")

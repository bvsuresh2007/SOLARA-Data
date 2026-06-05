"""
TallyExpenseParser — Extract expense entries from Tally Day Book XML files.

Uses Chart of Accounts for Group/SubGroup/Head hierarchy mapping.
Streams UTF-16 LE XML files of any size (tested up to 800 MB).

Usage:
    python -m scrapers.tools.tally_expense_parser.run <xml_path> [--output <xlsx_path>]

    # Multiple XMLs at once (auto-splits by month):
    python -m scrapers.tools.tally_expense_parser.run <xml1> <xml2> ... [--output-dir <dir>]

Examples:
    python -m scrapers.tools.tally_expense_parser.run "C:/Users/accou/Downloads/DayBook.xml Mar2024"
    python -m scrapers.tools.tally_expense_parser.run "C:/Downloads/APR-JUN.xml" --output-dir "C:/Downloads/Monthly/"
"""

import re
import os
import sys
import html
import argparse
from datetime import datetime
from collections import defaultdict, OrderedDict

import openpyxl
from openpyxl import Workbook

# ─── Chart of Accounts path (default) ────────────────────────────────────────
DEFAULT_COA_PATH = "C:/Users/accou/Downloads/Chart of accounts expenses.xlsx"
DEFAULT_MANUAL_PATH = "C:/Users/accou/Downloads/Mar24Expenses.xlsx"

CHUNK_BYTES = 2 * 1024 * 1024  # 2 MB

# ─── Section markers to SKIP as Group/SubGroup ───────────────────────────────
SKIP_MARKERS = {
    "expenses",
    "direct expenses",
    "selling & distribution expenses",
    "indirect expenses",
    "purchase accounts",
}

# Heads to EXCLUDE
EXCLUDE_HEADS = {"round off"}

# Purchase Accounts heads to EXCLUDE
PURCHASE_ACCOUNT_HEADS = {
    "liner charges", "handling charges", "custom duty", "social welfare charge",
    "import purchase", "custom charges", "cfs charges", "custom fine",
    "customs intrest", "india purchase", "purchase", "stock transfer inward",
    "mtr stock transfer inward", "reambersement of delivery order charges",
    "reambersement of igm filing charges", "reambersement of liner charges (srs)",
    "reambersement of manifestation charges", "reambersement of ocean freight",
    "reambersement of port charges", "reambersement of sims charges (srs)",
    "reambersement of telex release charges", "reambursement of con dam charges",
    "nanda international",
}

# ─── Party name normalization ────────────────────────────────────────────────
PARTY_NORMALIZATIONS = {
    "cod myntra b2c": "MYNTRA DESIGNS PVT LTD.",
    "myntra designs - (b2c)": "MYNTRA DESIGNS PVT LTD.",
    "snapmint received shopify": "Snapmint",
    "furniture and fixtures": "Fixed Assets",
    "tcs igst": "",
    "tcs dr reddys": "",
}

# Party overrides by expense head (when XML has no/wrong party)
PARTY_BY_HEAD = {
    "admin charges (wbb)": "EPFO",
    "audit fee": "Rajashekar & CO",
    "forex gain/loss": "Purchase",
}

# Post-processing aggregation across vouchers
AGGREGATE_HEADS = {"interest and penalties"}

# Head name aliases: (party_contains, xml_head_lower) → new_head
HEAD_ALIASES = {
    ("limechat", "marketing fees"): "Limechat Marketing",
    ("sharks", "promotional service"): "Sharks Promotional Service",
}

# ─── Regex patterns ──────────────────────────────────────────────────────────
VOUCHER_RE = re.compile(r'<VOUCHER\b[^>]*>(.*?)</VOUCHER>', re.DOTALL | re.IGNORECASE)
ENTRY_BLOCK_RE = re.compile(
    r'<(?:ALL)?LEDGERENTRIES\.LIST[^>]*>(.*?)</(?:ALL)?LEDGERENTRIES\.LIST>',
    re.DOTALL | re.IGNORECASE,
)


def decode_xml(s):
    if not s:
        return s
    return html.unescape(s).strip()


def extract_tag(text, tag):
    m = re.search(r'<' + tag + r'[^>]*>(.*?)</' + tag + r'>', text, re.DOTALL | re.IGNORECASE)
    return decode_xml(m.group(1)) if m else None


def parse_amount(s):
    if not s:
        return 0.0
    s = s.strip().replace(',', '')
    try:
        return float(s)
    except ValueError:
        return 0.0


def parse_date(s):
    try:
        return datetime.strptime(s, '%Y%m%d')
    except Exception:
        return None


def normalize_party(name):
    if not name:
        return name
    if name.startswith('Retailez Private Limited'):
        return 'Retailez Private Limited'
    norm = PARTY_NORMALIZATIONS.get(name.strip().lower())
    if norm is not None:
        return norm if norm else None
    return name


def is_liability_ledger(name):
    if not name:
        return True
    nu = name.upper()
    if 'PAYABLE' in nu:
        return True
    if nu.startswith('TDS') and len(nu) < 20:
        return True
    if nu in ('GST', 'RCM', 'CGST', 'SGST', 'IGST'):
        return True
    return False


def is_tax_entry(name):
    n = name.upper()
    return 'CGST' in n, 'SGST' in n, 'IGST' in n


# ─── Build mapping from Chart of Accounts ────────────────────────────────────
def build_coa_mapping(coa_path):
    """Parse Chart of Accounts Excel: bold = group headers, non-bold = leaf heads."""
    coa_mapping = {}
    wb = openpyxl.load_workbook(coa_path)
    ws = wb.active
    bold_window = []

    for row in ws.iter_rows():
        cell = row[0]
        val = cell.value
        if not val:
            continue
        val_str = str(val).strip()
        is_bold = bool(cell.font and cell.font.bold)

        if is_bold:
            bold_window.append(val_str)
            if len(bold_window) > 2:
                bold_window.pop(0)
        else:
            sg = None
            gr = None
            for b in reversed(bold_window):
                if b.lower() not in SKIP_MARKERS:
                    sg = b
                    break
            if sg is not None:
                found_sg = False
                for b in reversed(bold_window):
                    if b == sg and not found_sg:
                        found_sg = True
                        continue
                    if found_sg and b.lower() not in SKIP_MARKERS:
                        gr = b
                        break
            key = val_str.lower()
            if key not in coa_mapping:
                coa_mapping[key] = (gr, sg)

    wb.close()
    return coa_mapping


def build_manual_mapping(manual_path):
    """Extract Head → (Group, SubGroup) from a manually created reference Excel."""
    manual_mapping = {}
    if not manual_path or not os.path.exists(manual_path):
        return manual_mapping
    wb = openpyxl.load_workbook(manual_path)
    ws = wb.active
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        try:
            _, _, _, _, group, subgroup, head = row[:7]
        except (ValueError, TypeError):
            continue
        if head is None:
            continue
        key = str(head).strip().lower()
        if key not in manual_mapping:
            manual_mapping[key] = (group, subgroup)
    wb.close()
    return manual_mapping


# ─── Process a single voucher ────────────────────────────────────────────────
def process_voucher(v_text, combined_mapping, all_known_expense_heads):
    date_str = extract_tag(v_text, 'DATE')
    narration = extract_tag(v_text, 'NARRATION') or ''
    v_date = parse_date(date_str) if date_str else None

    entry_blocks = ENTRY_BLOCK_RE.findall(v_text)

    vch_party_name = extract_tag(v_text, 'PARTYNAME') or ''
    party_entry_name = None
    tax_cgst_total = 0.0
    tax_sgst_total = 0.0
    tax_igst_total = 0.0
    expense_entries = []

    for block in entry_blocks:
        ledger_name = extract_tag(block, 'LEDGERNAME') or ''
        amount_str = extract_tag(block, 'AMOUNT')
        is_party_str = extract_tag(block, 'ISPARTYLEDGER') or 'No'
        amount = parse_amount(amount_str)
        is_party = is_party_str.strip().lower() == 'yes'

        if is_party:
            if party_entry_name is None:
                party_entry_name = normalize_party(ledger_name)
            continue

        ic, is_, ig = is_tax_entry(ledger_name)
        if ic:
            tax_cgst_total += amount
        elif is_:
            tax_sgst_total += amount
        elif ig:
            tax_igst_total += amount
        else:
            lkey = ledger_name.strip().lower()
            if lkey in EXCLUDE_HEADS or lkey in PURCHASE_ACCOUNT_HEADS:
                continue
            if lkey in all_known_expense_heads:
                expense_entries.append((ledger_name, -amount))

    if not expense_entries:
        return []

    # Determine party name
    if vch_party_name and not is_liability_ledger(vch_party_name):
        party_name = normalize_party(vch_party_name)
    elif party_entry_name:
        party_name = party_entry_name
    else:
        party_name = None
        for lname, _ in expense_entries:
            lk = lname.strip().lower()
            if lk in PARTY_BY_HEAD:
                party_name = PARTY_BY_HEAD[lk]
                break
        if party_name is None:
            has_depreciation = any(
                lname.lower().startswith('depreciation') for lname, _ in expense_entries
            )
            party_name = 'Fixed Assets' if has_depreciation else None

    total_cgst = -tax_cgst_total
    total_sgst = -tax_sgst_total
    total_igst = -tax_igst_total

    if v_date:
        yr = v_date.year
        mo = v_date.month
        fy = f"FY {yr - 1}-{str(yr)[2:]}" if mo < 4 else f"FY {yr}-{str(yr + 1)[2:]}"
        month_label = v_date.strftime('%b') + str(v_date.year)[2:]
    else:
        fy = "Unknown"
        month_label = "Unknown"

    # Aggregate entries by Head within the same voucher
    agg = OrderedDict()
    for lname, amount in expense_entries:
        lkey = lname.strip().lower()
        if lkey in agg:
            agg[lkey] = (agg[lkey][0], agg[lkey][1] + amount)
        else:
            agg[lkey] = (lname, amount)

    total_taxable = sum(amt for _, amt in agg.values())

    # Apply head aliases (party-specific renames)
    party_lower = (party_name or '').lower()
    new_agg = OrderedDict()
    for lkey, (lname, amount) in agg.items():
        new_head = None
        for (party_contains, head_match), alias in HEAD_ALIASES.items():
            if party_contains in party_lower and head_match == lkey:
                new_head = alias
                break
        if new_head:
            new_key = new_head.strip().lower()
            new_agg[new_key] = (new_head, amount)
        else:
            new_agg[lkey] = (lname, amount)
    agg = new_agg

    # Skip zero-amount "Additions" entries
    agg = OrderedDict((k, v) for k, v in agg.items() if not (k == 'additions' and abs(v[1]) < 0.01))

    rows = []
    for lkey, (lname, amount) in agg.items():
        group, subgroup = combined_mapping.get(lkey, (None, None))

        if total_taxable != 0:
            ratio = amount / total_taxable
            cgst_alloc = total_cgst * ratio
            sgst_alloc = total_sgst * ratio
            igst_alloc = total_igst * ratio
        else:
            n = len(agg)
            cgst_alloc = total_cgst / n if n else 0
            sgst_alloc = total_sgst / n if n else 0
            igst_alloc = total_igst / n if n else 0

        cgst_out = round(cgst_alloc, 4) if abs(cgst_alloc) > 0.001 else None
        sgst_out = round(sgst_alloc, 4) if abs(sgst_alloc) > 0.001 else None
        igst_out = round(igst_alloc, 4) if abs(igst_alloc) > 0.001 else None

        rows.append({
            "fy": fy, "month": month_label, "date": v_date,
            "party": party_name, "group": group, "subgroup": subgroup,
            "head": lname, "narration": narration,
            "amount": round(amount, 2),
            "cgst": cgst_out, "sgst": sgst_out, "igst": igst_out,
        })

    return rows


# ─── Parse XML file (streaming) ──────────────────────────────────────────────
def parse_xml(xml_path, combined_mapping, all_known_expense_heads):
    """Stream-parse a Tally Day Book XML file and return expense rows."""
    generated_rows = []
    voucher_count = 0
    file_size = os.path.getsize(xml_path)
    print(f"  Parsing: {xml_path} ({file_size / 1024 / 1024:.1f} MB)")

    buffer = ""
    bytes_read = 0

    with open(xml_path, 'rb') as f:
        bom = f.read(2)
        assert bom == b'\xff\xfe', f"Unexpected BOM: {bom.hex()} — expected UTF-16 LE"

        while True:
            raw = f.read(CHUNK_BYTES)
            if not raw:
                break
            if len(raw) % 2 != 0:
                raw += f.read(1)
            buffer += raw.decode('utf-16-le', errors='replace')
            bytes_read += len(raw)

            last_end = 0
            for m in VOUCHER_RE.finditer(buffer):
                v_text = m.group(1)
                last_end = m.end()
                voucher_count += 1
                rows = process_voucher(v_text, combined_mapping, all_known_expense_heads)
                generated_rows.extend(rows)

            if last_end > 0:
                buffer = buffer[last_end:]

            pct = bytes_read / (file_size - 2) * 100
            sys.stdout.write(f"\r  Progress: {pct:.0f}%  vouchers={voucher_count}  rows={len(generated_rows)}   ")
            sys.stdout.flush()

        # Remaining buffer
        for m in VOUCHER_RE.finditer(buffer):
            v_text = m.group(1)
            voucher_count += 1
            rows = process_voucher(v_text, combined_mapping, all_known_expense_heads)
            generated_rows.extend(rows)

    print(f"\n  Done: {voucher_count} vouchers → {len(generated_rows)} expense rows")
    return generated_rows


# ─── Post-processing ─────────────────────────────────────────────────────────
def post_process(generated_rows):
    """Aggregate specific heads across vouchers on the same date."""
    if not AGGREGATE_HEADS:
        return generated_rows

    final_rows = []
    agg_bucket = {}
    for r in generated_rows:
        hk = r['head'].strip().lower()
        if hk in AGGREGATE_HEADS:
            key = (r['date'], hk)
            if key not in agg_bucket:
                agg_bucket[key] = []
            agg_bucket[key].append(r)
        else:
            final_rows.append(r)

    for (date, hk), rows_list in agg_bucket.items():
        merged = dict(rows_list[0])
        merged['amount'] = round(sum(r['amount'] for r in rows_list), 2)
        merged['cgst'] = sum(r['cgst'] or 0 for r in rows_list) or None
        merged['sgst'] = sum(r['sgst'] or 0 for r in rows_list) or None
        merged['igst'] = sum(r['igst'] or 0 for r in rows_list) or None
        for tax_key in ('cgst', 'sgst', 'igst'):
            if merged[tax_key] and abs(merged[tax_key]) < 0.001:
                merged[tax_key] = None
        if hk == 'interest and penalties':
            merged['party'] = 'GST'
        biggest = max(rows_list, key=lambda r: abs(r['amount']))
        merged['narration'] = biggest['narration']
        final_rows.append(merged)

    return final_rows


# ─── Write Excel output ─────────────────────────────────────────────────────
HEADERS = [
    "FY", "Month", "Date", "Party",
    "Tally Expense Group", "Tally Sub Group", "Tally Head",
    "Narration", "Taxable Amount", "Cgst", "Sgst", "Igst",
]


def write_excel(rows, output_path):
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Sheet1")
    ws.append(HEADERS)
    for r in rows:
        ws.append([
            r["fy"], r["month"], r["date"], r["party"],
            r["group"], r["subgroup"], r["head"],
            r["narration"], r["amount"], r["cgst"], r["sgst"], r["igst"],
        ])
    wb.save(output_path)
    return len(rows)


def write_monthly_excels(rows, output_dir):
    """Split rows by month and write one Excel per month."""
    os.makedirs(output_dir, exist_ok=True)
    by_month = OrderedDict()
    for r in rows:
        key = r['month']  # e.g. "Mar24", "Apr24"
        if key not in by_month:
            by_month[key] = []
        by_month[key].append(r)

    files = []
    for month_label, month_rows in by_month.items():
        # Derive a filename like "Expenses_Mar24.xlsx"
        fname = f"Expenses_{month_label}.xlsx"
        fpath = os.path.join(output_dir, fname)
        count = write_excel(month_rows, fpath)
        files.append((fpath, count))
        print(f"  {fname}: {count} rows")
    return files


# ─── Main ────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description="TallyExpenseParser: Extract expenses from Tally Day Book XML",
    )
    parser.add_argument("xml_paths", nargs="+", help="Path(s) to Tally Day Book XML file(s)")
    parser.add_argument("--coa", default=DEFAULT_COA_PATH, help="Path to Chart of Accounts Excel")
    parser.add_argument("--manual", default=DEFAULT_MANUAL_PATH, help="Path to manual reference Excel (for mapping)")
    parser.add_argument("--output", help="Output Excel path (single file)")
    parser.add_argument("--output-dir", help="Output directory (one Excel per month)")
    parser.add_argument("--no-manual", action="store_true", help="Skip manual mapping, use only CoA")

    args = parser.parse_args()

    # Build mappings
    print("=== Building mappings ===")
    coa_mapping = build_coa_mapping(args.coa)
    print(f"  CoA mapping: {len(coa_mapping)} heads")

    combined_mapping = dict(coa_mapping)
    if not args.no_manual:
        manual_mapping = build_manual_mapping(args.manual)
        combined_mapping.update(manual_mapping)
        print(f"  Manual mapping: {len(manual_mapping)} heads")
    print(f"  Combined: {len(combined_mapping)} heads")

    all_known_expense_heads = set(combined_mapping.keys())

    # Parse all XMLs
    all_rows = []
    for xml_path in args.xml_paths:
        print(f"\n=== Parsing XML ===")
        rows = parse_xml(xml_path, combined_mapping, all_known_expense_heads)
        all_rows.extend(rows)

    # Post-process
    all_rows = post_process(all_rows)
    print(f"\n  Total expense rows after post-processing: {len(all_rows)}")

    # Write output
    if args.output_dir:
        print(f"\n=== Writing monthly Excels to {args.output_dir} ===")
        files = write_monthly_excels(all_rows, args.output_dir)
        print(f"\n  Total files: {len(files)}")
    elif args.output:
        print(f"\n=== Writing output ===")
        count = write_excel(all_rows, args.output)
        print(f"  Saved: {args.output} ({count} rows)")
    else:
        # Default: monthly files in same dir as first XML
        output_dir = os.path.join(os.path.dirname(args.xml_paths[0]), "Expenses_Monthly")
        print(f"\n=== Writing monthly Excels to {output_dir} ===")
        files = write_monthly_excels(all_rows, output_dir)
        print(f"\n  Total files: {len(files)}")


if __name__ == "__main__":
    main()
